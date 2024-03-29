import customtkinter as ctk
import math
import sqlite3
from PIL import Image
from edit import Edit

import pandas as pd
import numpy as np
import time


from pandas import ExcelWriter, ExcelFile
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

# from ls import loading_page as ls
# import threading


class ScrollableCheckBoxFrame(ctk.CTkFrame):
    def __init__(self, master, item_list, command=None, **kwargs):
        super().__init__(master, **kwargs)

        self.command = command
        self.checkbox_list = []
        for i, item in enumerate(item_list):
            self.add_item(item)

    def add_item(self, item):
        checkbox = ctk.CTkCheckBox(self, text=item)
        if self.command is not None:
            checkbox.configure(command=self.command)
        checkbox.grid(row=0, column=len(self.checkbox_list), pady=(0, 10))
        self.checkbox_list.append(checkbox)

    def remove_item(self, item):
        for checkbox in self.checkbox_list:
            if item == checkbox.cget("text"):
                checkbox.destroy()
                self.checkbox_list.remove(checkbox)
                return

    def get_checked_items(self):
        return [
            checkbox.cget("text")
            for checkbox in self.checkbox_list
            if checkbox.get() == 1
        ]


class main(ctk.CTk):
    ctk.set_appearance_mode("light")

    def __init__(self):
        super().__init__()

        self.wins = 0

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        self.navigation_frame = ctk.CTkFrame(self, corner_radius=0)
        self.navigation_frame.grid(row=0, column=0, sticky="nsew")
        # self.navigation_frame.grid_rowconfigure(9, weight=1)

        self.calcFrame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.calcFrame.grid_columnconfigure(0, weight=1)
        self.calcFrame.grid_columnconfigure(1, weight=1)

        self.sheetFrame = ctk.CTkScrollableFrame(
            self, corner_radius=0, fg_color="transparent"
        )
        # self.sheetFrame.grid_columnconfigure(0, weight=1)
        # self.sheetFrame.grid_columnconfigure(1, weight=1)

        # self.calcFrame.grid(row=0, column=1, sticky="nsew")

        # creating databases
        self.conn = sqlite3.connect("data.db")
        self.cur = self.conn.cursor()
        self.cur.execute(
            """CREATE TABLE IF NOT EXISTS diameter (
                id INTEGER PRIMARY KEY,
                pipe_diameter REAL NOT NULL,
                min_wall_thickness REAL NOT NULL,
                mold_diameter REAL NOT NULL, 
                mold_optimal_temperature REAL);"""
        )

        self.cur.execute(
            """CREATE TABLE IF NOT EXISTS ppwt (
                                        id integer PRIMARY KEY,
                                        pp_diameter REAL NOT NULL,
                                        weight REAL NOT NULL
                                    );"""
        )

        self.cur.execute(
            """ CREATE TABLE IF NOT EXISTS rawMaterial (
                                        id INTEGER PRIMARY KEY,
                                        profile TEXT NOT NULL,
                                        density REAL NOT NULL,
                                        elastic_modulus REAL NOT NULL,
                                        shrinkage REAL NOT NULL
                                    ); """
        )
        self.cur.execute(
            """ CREATE TABLE IF NOT EXISTS flatDie (
                                        id INTEGER PRIMARY KEY,
                                        profile TEXT NOT NULL,
                                        pitch REAL NOT NULL,
                                        pitch_factor REAL NOT NULL,
                                        thickness REAL NOT NULL
                                    ); """
        )
        self.cur.execute(
            """ CREATE TABLE IF NOT EXISTS claddingDie (
                                        id INTEGER PRIMARY KEY,
                                        profile TEXT NOT NULL,
                                        pp_diameter REAL NOT NULL,
                                        pp_thickness REAL NOT NULL
                                    ); """
        )
        self.conn.commit()

        # app
        self.title("Stiffness calculator")
        self.W = 1300
        self.H = 800

        self.geometry(
            str(self.W)
            + "x"
            + str(self.H)
            + "+"
            + str(int((self.winfo_screenwidth() - self.W) / 2))
            + "+"
            + str(int((self.winfo_screenheight() - self.H) / 2))
        )

        # Creaeting Frames

        # labels

        self.tablesLabel = ctk.CTkLabel(
            self.navigation_frame,
            fg_color="transparent",
            text="Tables",
            compound="left",
            font=ctk.CTkFont(size=15, weight="bold"),
        )
        self.tablesLabel.grid(row=0, column=0, pady=5)

        self.profilesLabel = ctk.CTkLabel(
            self.navigation_frame,
            fg_color="transparent",
            text="Profiles",
            compound="left",
            font=ctk.CTkFont(size=15, weight="bold"),
        )
        self.profilesLabel.grid(row=5, column=0, pady=5)

        self.calculatorLabel = ctk.CTkLabel(
            self.navigation_frame,
            fg_color="transparent",
            text="Calculators",
            compound="left",
            font=ctk.CTkFont(size=15, weight="bold"),
        )
        self.calculatorLabel.grid(row=9, column=0, pady=5)

        #

        #

        # buttons

        self.ppwtTable = ctk.CTkButton(
            self.navigation_frame,
            corner_radius=0,
            height=40,
            border_spacing=10,
            fg_color="transparent",
            text_color=("gray10", "gray90"),
            hover_color=("gray70", "gray30"),
            anchor="w",
            text="pp weight",
            command=lambda: self.select_frame_by_name("ppwt"),
        )
        self.ppwtTable.grid(row=1, column=0, pady=5, padx=5)

        self.diameterTable = ctk.CTkButton(
            self.navigation_frame,
            corner_radius=0,
            height=40,
            border_spacing=10,
            fg_color="transparent",
            text_color=("gray10", "gray90"),
            hover_color=("gray70", "gray30"),
            anchor="w",
            text="Diameter",
            command=lambda: self.select_frame_by_name("diameter"),
        )
        self.diameterTable.grid(row=2, column=0, pady=5, padx=2)

        self.materialProfileB = ctk.CTkButton(
            self.navigation_frame,
            corner_radius=0,
            height=40,
            border_spacing=10,
            fg_color="transparent",
            text_color=("gray10", "gray90"),
            hover_color=("gray70", "gray30"),
            anchor="w",
            text="material",
            command=lambda: self.select_frame_by_name("rawMaterial"),
        )

        self.materialProfileB.grid(row=6, column=0, pady=5, padx=2)
        self.flatDiaProfileB = ctk.CTkButton(
            self.navigation_frame,
            corner_radius=0,
            height=40,
            border_spacing=10,
            fg_color="transparent",
            text_color=("gray10", "gray90"),
            hover_color=("gray70", "gray30"),
            anchor="w",
            text="flat die",
            command=lambda: self.select_frame_by_name("flatDie"),
        )
        self.flatDiaProfileB.grid(row=7, column=0, pady=5, padx=2)
        self.claddingDieProfileB = ctk.CTkButton(
            self.navigation_frame,
            corner_radius=0,
            height=40,
            border_spacing=10,
            fg_color="transparent",
            text_color=("gray10", "gray90"),
            hover_color=("gray70", "gray30"),
            anchor="w",
            text="cladding die",
            command=lambda: self.select_frame_by_name("claddingDie"),
        )
        self.claddingDieProfileB.grid(row=8, column=0, pady=5, padx=2)

        self.calculatorB = ctk.CTkButton(
            self.navigation_frame,
            corner_radius=0,
            height=40,
            border_spacing=10,
            fg_color="transparent",
            text_color=("gray10", "gray90"),
            hover_color=("gray70", "gray30"),
            text="stiffness",
            anchor="w",
            command=lambda: self.select_frame_by_name("calculator"),
        )
        self.calculatorB.grid(row=10, column=0, pady=5, padx=2)

        self.sheetB = ctk.CTkButton(
            self.navigation_frame,
            corner_radius=0,
            height=40,
            border_spacing=10,
            fg_color="transparent",
            text_color=("gray10", "gray90"),
            hover_color=("gray70", "gray30"),
            text="Export Sheet",
            anchor="w",
            command=lambda: self.select_frame_by_name("sheet"),
        )
        self.sheetB.grid(row=11, column=0, pady=5, padx=2)

        self.scrollable_checkbox_frame = ScrollableCheckBoxFrame(
            master=self.sheetFrame,
            width=600,
            height=5,
            command=self.buttonCommand,
            item_list=["2", "4", "6.3", "8", "12.5", "16"],
            bg_color="transparent",
            fg_color="transparent",
        )

        self.scrollable_checkbox_frame.grid(
            row=0, column=1, padx=15, pady=15, columnspan=10
        )

        #

        #

        #
        # self.setResize(False)
        # self.resizable(False, False)
        #

        #

        self.modeNames = ["Select Mode", "VW", "PR", "SQ", "OP"]
        self.elastic_modulus = -1
        self.density = -1
        self.shrinkage = -1

        self.pipe_length = -1
        self.pd = -1
        self.p = -1
        self.s1 = -1  # wall_thickness
        self.ppd = -1
        self.s4 = -1  # pp_thickness

        self.pp_dist = -1
        self.Sn = -1
        self.W0 = -1
        self.W3 = -1
        self.W4 = -1

        # Creaeting Frames

        self.Frame = ctk.CTkFrame(self)
        self.Headers = []
        self.RowsData = []
        self.NumRows = int
        self.Entries = []

        self.rawMaterialFrame = ctk.CTkFrame(
            self.calcFrame, width=(self.W * 0.872), height=(self.H / 11)
        )
        self.rawMaterialFrame.grid_propagate(False)
        self.rawMaterialFrame.grid(
            padx=5, pady=10, sticky="n", row=0, column=1, columnspan=2
        )

        self.productionFrame = ctk.CTkFrame(
            self.calcFrame, width=(self.W * 0.872), height=(self.H / 4)
        )
        self.productionFrame.grid_propagate(False)
        self.productionFrame.grid(
            padx=5, pady=10, sticky="n", row=1, column=1, columnspan=2
        )

        self.resultsFrame = ctk.CTkFrame(
            self.calcFrame, width=(self.W * 0.872), height=(self.H / 4) + 10
        )
        self.resultsFrame.grid_propagate(False)
        self.resultsFrame.grid(
            padx=5, pady=10, sticky="n", row=4, column=1, columnspan=2, rowspan=2
        )

        self.ls = ctk.CTkFrame(
            self.resultsFrame, bg_color="transparent", fg_color="transparent"
        )
        self.ls2 = ctk.CTkFrame(
            self.sheetFrame, bg_color="transparent", fg_color="transparent"
        )
        self.loading_label = ctk.CTkLabel(
            self.ls,
            text="Optimizing...",
            font=("Arial", 14),
            bg_color="transparent",
            fg_color="transparent",
        )

        self.loading_label.grid(
            row=0,
            column=0,
            rowspan=2,
        )

        self.loading_label2 = ctk.CTkLabel(
            self.ls2,
            text="Optimizing...",
            font=("Arial", 14),
            bg_color="transparent",
            fg_color="transparent",
        )

        self.loading_label2.grid(
            row=0,
            column=0,
            rowspan=2,
        )

        self.modeFrame = ctk.CTkFrame(
            self.rawMaterialFrame,
            width=self.W * (1 - 0.72222),
            height=(self.H / 20),
            fg_color="transparent",
        )

        self.modeFrame.grid(
            padx=5, pady=10, sticky="ne", row=0, column=0, columnspan=3, rowspan=2
        )

        #

        #

        #

        # self.modes = ctk.CTkOptionMenu(
        #     self.modeFrame,
        #     values=self.modeNames,
        #     command=self.modeCommand,
        #     corner_radius=0,
        #     fg_color="#4eb56b",
        #     button_color="#2b6e3e",
        #     button_hover_color="#235731",
        # )
        self.mode = ""

        self.radio_var = ctk.IntVar(value=0)
        self.PRBtn = ctk.CTkRadioButton(
            self.modeFrame,
            text="PR",
            command=lambda: self.modeCommand("PR"),
            radiobutton_width=20,
            radiobutton_height=19,
            variable=self.radio_var,
            value=1,
        )
        self.PRBtn.grid(row=0, column=1, sticky="w", columnspan=1)
        self.VWBtn = ctk.CTkRadioButton(
            self.modeFrame,
            text="VW",
            command=lambda: self.modeCommand("VW"),
            radiobutton_width=20,
            radiobutton_height=19,
            variable=self.radio_var,
            value=2,
        )
        self.VWBtn.grid(row=0, column=2, sticky="w", columnspan=1)
        self.modesLabel = ctk.CTkLabel(
            self.modeFrame,
            text="Mode:       ",
            fg_color="transparent",
            anchor="w",
        )
        self.modesLabel.grid(row=0, column=0, pady=2, sticky="w")

        items = self.cur.execute("SELECT profile FROM rawMaterial").fetchall()
        items = [item[0] for item in items]

        self.materialprofile = ctk.CTkOptionMenu(
            self.rawMaterialFrame,
            values=list(items),
            anchor="w",
            command=self.materialCommand,
            corner_radius=0,
            dynamic_resizing=False,
            fg_color="#4eb56b",
            button_color="#2b6e3e",
            button_hover_color="#235731",
        )

        self.materialprofile.set("Material Profiles")

        items = self.cur.execute("SELECT profile FROM flatDie").fetchall()
        items = [item[0] for item in items]

        self.flatDieProfile = ctk.CTkOptionMenu(
            self.productionFrame,
            values=list(items),
            command=self.dieCommand,
            anchor="w",
            corner_radius=0,
            dynamic_resizing=False,
            fg_color="#4eb56b",
            button_color="#2b6e3e",
            button_hover_color="#235731",
            width=165,
        )

        self.flatDieProfile.set("Flat die profiles")

        items = self.cur.execute("SELECT profile FROM claddingDie").fetchall()
        items = [item[0] for item in items]

        self.claddingDieProfile = ctk.CTkOptionMenu(
            self.productionFrame,
            values=list(items),
            command=self.claddingCommand,
            anchor="w",
            corner_radius=0,
            dynamic_resizing=False,
            width=165,
            fg_color="#4eb56b",
            button_color="#2b6e3e",
            button_hover_color="#235731",
        )
        self.claddingDieProfile.set("Cladding die profiles")

        #

        #

        #

        #

        #

        # raw material
        self.rawMaterialError = ctk.CTkLabel(
            self.rawMaterialFrame, text="", text_color="red"
        )
        self.rawMaterialError.grid(
            row=0, column=3, padx=10, pady=5, columnspan=3, sticky="w"
        )
        self.SnL = ctk.CTkLabel(
            self.resultsFrame,
        )
        self.W0L = ctk.CTkLabel(
            self.resultsFrame,
        )
        self.WC = ctk.CTkLabel(
            self.resultsFrame,
        )
        self.WP = ctk.CTkLabel(
            self.resultsFrame,
        )
        self.rawMaterilaLabel = ctk.CTkLabel(
            self.rawMaterialFrame,
            text="Raw Material:",
            fg_color="transparent",
            anchor="w",
            width=10,
        )

        self.rawMaterilaLabel.grid(
            row=1, column=0, padx=5, pady=2, columnspan=2, sticky="w"
        )

        self.materialNameLabel = ctk.CTkLabel(
            self.rawMaterialFrame,
            text="Profile Name",
            fg_color="transparent",
            anchor="w",
            width=10,
        )
        self.materialNameEntry = ctk.CTkEntry(
            self.rawMaterialFrame, placeholder_text="", width=100
        )

        self.densityLabel = ctk.CTkLabel(
            self.rawMaterialFrame,
            text="Density",
            fg_color="transparent",
            anchor="w",
            width=10,
        )
        self.densityLabel.grid_forget()
        self.densityEntry = ctk.CTkEntry(
            self.rawMaterialFrame,
            placeholder_text="",
            width=50,
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )
        self.densityUnitLabel = ctk.CTkLabel(
            self.rawMaterialFrame,
            text="(g/cm³)",
            fg_color="transparent",
        )
        self.densityEntry.grid_forget()

        self.elasticLabel = ctk.CTkLabel(
            self.rawMaterialFrame, fg_color="transparent", text="Elastic Modulus"
        )
        self.elasticLabel.grid_forget()
        self.elasticEntry = ctk.CTkEntry(
            self.rawMaterialFrame,
            placeholder_text="",
            width=50,
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )

        self.elasticUnitLabel = ctk.CTkLabel(
            self.rawMaterialFrame, fg_color="transparent", text="(Mpa)"
        )
        self.elasticEntry.grid_forget()

        self.shrinkageLabel = ctk.CTkLabel(
            self.rawMaterialFrame, fg_color="transparent", text="Shrinkage Rate"
        )
        self.shrinkageLabel.grid_forget()
        self.shrinkageUnitLabel = ctk.CTkLabel(
            self.rawMaterialFrame, fg_color="transparent", text="(%)"
        )
        self.shrinkageEntry = ctk.CTkEntry(
            self.rawMaterialFrame,
            placeholder_text="",
            width=50,
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )
        self.shrinkageEntry.grid_forget()

        #

        #

        #

        #

        #

        # production
        self.productionError = ctk.CTkLabel(
            self.productionFrame, text="", text_color="red"
        )
        self.productionError.grid(
            row=0, column=1, padx=10, pady=5, columnspan=3, sticky="w"
        )
        self.productionName = ctk.CTkLabel(
            self.productionFrame,
            fg_color="transparent",
            text="Production:",
        )
        self.productionName.grid(
            row=0, column=0, padx=5, pady=5, columnspan=2, sticky="w"
        )

        self.dieNameLabel = ctk.CTkLabel(
            self.productionFrame,
            text="Profile Name",
            fg_color="transparent",
            anchor="w",
            width=10,
        )
        self.dieNameEntry = ctk.CTkEntry(
            self.productionFrame, placeholder_text="", width=150
        )

        self.claddingNameLabel = ctk.CTkLabel(
            self.productionFrame,
            text="Profile Name",
            fg_color="transparent",
            anchor="w",
            width=10,
        )
        self.claddingNameEntry = ctk.CTkEntry(
            self.productionFrame, placeholder_text="", width=150
        )

        self.pipeLengthLabel = ctk.CTkLabel(
            self.productionFrame, fg_color="transparent", text="Pipe Length"
        )
        self.pipeLengthLabel.grid_forget()
        self.pipeLengthUnitLabel = ctk.CTkLabel(
            self.productionFrame, fg_color="transparent", text="(mm)"
        )
        self.pipeLengthEnry = ctk.CTkEntry(
            self.productionFrame, placeholder_text="", width=70
        )
        self.pipeLengthEnry.grid_forget()

        self.pipeDiameterLabel = ctk.CTkLabel(
            self.productionFrame, fg_color="transparent", text="Pipe Diameter"
        )
        self.pipeDiameterLabel.grid_forget()
        self.pipeDiameterUnitLabel = ctk.CTkLabel(
            self.productionFrame, fg_color="transparent", text="(mm)"
        )
        self.pipeDiameterEnry = ctk.CTkEntry(
            self.productionFrame, placeholder_text="", width=70
        )
        self.pipeDiameterEnry.grid_forget()

        self.pitchLabel = ctk.CTkLabel(
            self.productionFrame, fg_color="transparent", text="Flat die"
        )
        self.pitchUnitLabel = ctk.CTkLabel(
            self.productionFrame, fg_color="transparent", text="(mm)"
        )
        self.pitchUnitLabel.grid_forget()
        self.pitchEntry = ctk.CTkEntry(
            self.productionFrame,
            placeholder_text="",
            width=70,
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )
        self.pitchEntry.grid_forget()

        self.pitchEntry.bind("<Return>", command=self.calcPitch)

        self.pitchFactorLabel = ctk.CTkLabel(
            self.productionFrame, fg_color="transparent", text="Flat Die Factor"
        )
        self.pitchFactorUnitLabel = ctk.CTkLabel(
            self.productionFrame, fg_color="transparent", text="(mm)"
        )
        self.pitchFactorUnitLabel.grid_forget()
        self.pitchFactorEntry = ctk.CTkEntry(
            self.productionFrame, placeholder_text="", width=70, state="normal"
        )
        self.pitchFactorEntry.grid_forget()

        self.pitchFactorEntry.bind("<Return>", command=self.calcPitch)

        self.finalPitchLabel = ctk.CTkLabel(
            self.productionFrame, fg_color="transparent", text="Pitch"
        )
        self.finalPitchUnitLabel = ctk.CTkLabel(
            self.productionFrame, fg_color="transparent", text="(mm)"
        )
        self.finalPitchUnitLabel.grid_forget()
        self.finalPitchEntry = ctk.CTkEntry(
            self.productionFrame,
            placeholder_text="",
            width=70,
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )
        self.finalPitchEntry.grid_forget()

        self.WallThicknessLabel = ctk.CTkLabel(
            self.productionFrame, fg_color="transparent", text="Wall Thickness"
        )
        self.WallThicknessLabel.grid_forget()
        self.WallThicknessUnitLabel = ctk.CTkLabel(
            self.productionFrame, fg_color="transparent", text="(mm)"
        )
        self.WallThicknessEntry = ctk.CTkEntry(
            self.productionFrame,
            placeholder_text="",
            width=70,
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )
        self.WallThicknessEntry.grid_forget()

        self.PPDiameterLabel = ctk.CTkLabel(
            self.productionFrame, fg_color="transparent", text="PP Diameter"
        )
        self.PPDiameterLabel.grid_forget()
        self.PPDiameterUnitLabel = ctk.CTkLabel(
            self.productionFrame, fg_color="transparent", text="(mm)"
        )
        self.PPDiameterEntry = ctk.CTkEntry(
            self.productionFrame,
            placeholder_text="",
            width=70,
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )
        self.PPDiameterEntry.grid_forget()

        self.PPFilmThicknessLabel = ctk.CTkLabel(
            self.productionFrame, fg_color="transparent", text="PP Film Thickness"
        )
        self.PPFilmThicknessLabel.grid_forget()
        self.PPFilmThicknessUnitLabel = ctk.CTkLabel(
            self.productionFrame, fg_color="transparent", text="(mm)"
        )
        self.PPFilmThicknessEntry = ctk.CTkEntry(
            self.productionFrame,
            placeholder_text="",
            width=70,
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )
        self.PPFilmThicknessEntry.grid_forget()

        #

        #

        #

        #

        #

        self.MachineTabs = ctk.CTkTabview(
            self.calcFrame,
            width=(self.W * 0.872),
            height=(self.H / 4) + 40,
            segmented_button_selected_color="#5696b0",
        )
        self.MachineTabs.grid(
            padx=5, pady=10, sticky="n", row=2, column=1, columnspan=2, rowspan=2
        )
        self.MachineTabs.grid_propagate(False)
        self.MachineTabs.add("Extruder")
        self.MachineTabs.add("Machine")
        self.MachineTabs.add("Trolley Position")
        self.machineLimitsError = ctk.CTkLabel(
            self.MachineTabs.tab("Machine"), text="", text_color="red"
        )
        self.machineLimitsError.grid(
            row=1, column=7, padx=10, pady=5, sticky="nw", columnspan=6
        )

        self.maxFactorLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"),
            fg_color="transparent",
            text="Production Factor",
        )
        self.maxFactorUnitLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"), fg_color="transparent", text="(%)"
        )
        self.maxFactorEntry = ctk.CTkEntry(
            self.MachineTabs.tab("Extruder"), placeholder_text="", width=70
        )
        self.maxFactorEntry.delete(0, ctk.END)
        self.maxFactorEntry.insert(0, "80")

        self.flat75MaxLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"),
            fg_color="transparent",
            text="Flat Extruder 75 Max",
        )
        self.flat75MaxUnitLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"), fg_color="transparent", text="(kg/hr)"
        )
        self.flat75MaxEntry = ctk.CTkEntry(
            self.MachineTabs.tab("Extruder"), placeholder_text="", width=70
        )

        self.flat75MaxEntry.delete(0, ctk.END)
        self.flat75MaxEntry.insert(0, "420")

        self.flat45MaxLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"),
            fg_color="transparent",
            text="Flat Extruder 45 Max",
        )
        self.flat45MaxUnitLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"), fg_color="transparent", text="(kg/hr)"
        )
        self.flat45MaxEntry = ctk.CTkEntry(
            self.MachineTabs.tab("Extruder"), placeholder_text="", width=70
        )
        self.flat45MaxEntry.delete(0, ctk.END)
        self.flat45MaxEntry.insert(0, "90")

        self.cladding75MaxLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"),
            fg_color="transparent",
            text="Cladding Extruder 75 Max",
        )
        self.cladding75MaxUnitLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"), fg_color="transparent", text="(kg/hr)"
        )
        self.cladding75MaxEntry = ctk.CTkEntry(
            self.MachineTabs.tab("Extruder"), placeholder_text="", width=70
        )
        self.cladding75MaxEntry.delete(0, ctk.END)
        self.cladding75MaxEntry.insert(0, "420")
        self.cladding45MaxLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"),
            fg_color="transparent",
            text="Cladding Extruder 45 Max",
        )
        self.cladding45MaxUnitLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"), fg_color="transparent", text="(kg/hr)"
        )
        self.cladding45MaxEntry = ctk.CTkEntry(
            self.MachineTabs.tab("Extruder"), placeholder_text="", width=70
        )
        self.cladding45MaxEntry.delete(0, ctk.END)
        self.cladding45MaxEntry.insert(0, "90")
        self.MoldSpeedLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Machine"), fg_color="transparent", text="Mold Speed"
        )
        self.MoldSpeedUnitLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Machine"), fg_color="transparent", text="(mm/min)"
        )
        self.MoldSpeedLabel.grid_forget()

        self.MoldSpeedEntry = ctk.CTkEntry(
            self.MachineTabs.tab("Machine"), placeholder_text="", width=70
        )
        self.MoldSpeedEntry.delete(0, ctk.END)
        self.MoldSpeedEntry.insert(0, "3850")

        self.MoldSpeedEntry.bind("<Return>", command=self.calcSpeeds)
        self.ppSpeedFractionLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Machine"),
            fg_color="transparent",
            text="PP Speed Fraction",
        )
        self.ppSpeedFractionUnitLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Machine"), fg_color="transparent", text="(%)"
        )
        self.ppSpeedFractionLabel.grid_forget()
        self.ppSpeedFractionEntry = ctk.CTkEntry(
            self.MachineTabs.tab("Machine"), placeholder_text="", width=70
        )
        self.ppSpeedFractionEntry.delete(0, ctk.END)
        self.ppSpeedFractionEntry.insert(0, "95")

        self.ppSpeedFractionEntry.bind("<Return>", command=self.calcSpeeds)
        self.ppSpeedLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Machine"), fg_color="transparent", text="PP Speed"
        )
        self.ppSpeedUnitLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Machine"), fg_color="transparent", text="(mm/min)"
        )
        self.ppSpeedLabel.grid_forget()
        self.ppSpeedEntry = ctk.CTkEntry(
            self.MachineTabs.tab("Machine"),
            placeholder_text="",
            width=70,
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )
        self.trolleySpeedLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Machine"),
            fg_color="transparent",
            text="Trolley Speed",
        )
        self.trolleySpeedUnitLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Machine"), fg_color="transparent", text="(mm/min)"
        )
        self.trolleySpeedLabel.grid_forget()
        self.trolleySpeedEntry = ctk.CTkEntry(
            self.MachineTabs.tab("Machine"),
            placeholder_text="",
            width=70,
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )

        self.FlatExtruder75Label = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"),
            fg_color="transparent",
            text="Flat Extruder 75",
        )
        self.FlatExtruder75UnitLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"), fg_color="transparent", text="(%)"
        )
        self.FlatExtruder75Label.grid_forget()
        self.FlatExtruder75Entry = ctk.CTkEntry(
            self.MachineTabs.tab("Extruder"), placeholder_text="", width=70
        )
        self.FlatExtruder75Entry.delete(0, ctk.END)
        self.FlatExtruder75Entry.insert(0, "83")

        self.FlatExtruder75Entry.bind("<Return>", command=self.f45)

        self.FlatExtruder45Label = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"),
            fg_color="transparent",
            text="Flat Extruder 45",
        )
        self.FlatExtruder45UnitLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"),
            fg_color="transparent",
            text="(%)",
        )
        self.FlatExtruder45Label.grid_forget()
        self.FlatExtruder45Entry = ctk.CTkEntry(
            self.MachineTabs.tab("Extruder"),
            placeholder_text="",
            width=70,
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )

        self.CladdingExtruder75Label = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"),
            fg_color="transparent",
            text="Cladding Extruder 75",
        )
        self.CladdingExtruder75UnitLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"), fg_color="transparent", text="(%)"
        )
        self.CladdingExtruder75Label.grid_forget()
        self.CladdingExtruder75Entry = ctk.CTkEntry(
            self.MachineTabs.tab("Extruder"), placeholder_text="", width=70
        )
        self.CladdingExtruder75Entry.delete(0, ctk.END)
        self.CladdingExtruder75Entry.insert(0, "83")

        self.CladdingExtruder75Entry.bind("<Return>", command=self.c45)

        self.CladdingExtruder45Label = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"),
            fg_color="transparent",
            text="Cladding Extruder 45",
        )
        self.CladdingExtruder45UnitLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"), fg_color="transparent", text="(%)"
        )
        self.CladdingExtruder45Label.grid_forget()
        self.CladdingExtruder45Entry = ctk.CTkEntry(
            self.MachineTabs.tab("Extruder"),
            placeholder_text="",
            width=70,
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )

        self.FlatExtruder75LabelF = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"),
            fg_color="transparent",
            text="Flat Extruder 75",
        )
        self.FlatExtruder75UnitLabelF = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"), fg_color="transparent", text="(kg/hr)"
        )
        self.FlatExtruder75LabelF.grid_forget()
        self.FlatExtruder75EntryF = ctk.CTkEntry(
            self.MachineTabs.tab("Extruder"),
            placeholder_text="",
            width=70,
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )

        self.FlatExtruder45LabelF = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"),
            fg_color="transparent",
            text="Flat Extruder 45",
        )
        self.FlatExtruder45UnitLabelF = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"),
            fg_color="transparent",
            text="(kg/hr)",
        )
        self.FlatExtruder45LabelF.grid_forget()
        self.FlatExtruder45EntryF = ctk.CTkEntry(
            self.MachineTabs.tab("Extruder"),
            placeholder_text="",
            width=70,
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )

        self.CladdingExtruder75LabelF = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"),
            fg_color="transparent",
            text="Cladding Extruder 75",
        )
        self.CladdingExtruder75UnitLabelF = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"), fg_color="transparent", text="(kg/hr)"
        )
        self.CladdingExtruder75LabelF.grid_forget()
        self.CladdingExtruder75EntryF = ctk.CTkEntry(
            self.MachineTabs.tab("Extruder"),
            placeholder_text="",
            width=70,
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )

        self.CladdingExtruder45LabelF = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"),
            fg_color="transparent",
            text="Cladding Extruder 45",
        )
        self.CladdingExtruder45UnitLabelF = ctk.CTkLabel(
            self.MachineTabs.tab("Extruder"), fg_color="transparent", text="(kg/hr)"
        )
        self.CladdingExtruder45LabelF.grid_forget()
        self.CladdingExtruder45EntryF = ctk.CTkEntry(
            self.MachineTabs.tab("Extruder"),
            placeholder_text="",
            width=70,
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )

        self.MouthStartLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Trolley Position"),
            fg_color="transparent",
            text="Mouth Start \n(mm)",
        )
        self.MouthStartEntry = ctk.CTkEntry(
            self.MachineTabs.tab("Trolley Position"), placeholder_text="", width=70
        )
        self.MouthEndLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Trolley Position"),
            fg_color="transparent",
            text="Mouth End \n(mm)",
        )
        self.MouthEndEntry = ctk.CTkEntry(
            self.MachineTabs.tab("Trolley Position"), placeholder_text="", width=70
        )
        self.CarriageReturnPositionLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Trolley Position"),
            fg_color="transparent",
            text="Carriage Return \nPosition (mm)",
        )
        self.CarriageReturnPositionEntry = ctk.CTkEntry(
            self.MachineTabs.tab("Trolley Position"), placeholder_text="", width=70
        )
        self.CarriageRotationsLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Trolley Position"),
            fg_color="transparent",
            text="Mold Rotations",
        )
        self.CarriageRotationsEntry = ctk.CTkEntry(
            self.MachineTabs.tab("Trolley Position"), placeholder_text="", width=70
        )
        self.CarriageReturnDelayLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Trolley Position"),
            fg_color="transparent",
            text="Carriage Return \nDelay (s)",
        )
        self.CarriageReturnDelayEntry = ctk.CTkEntry(
            self.MachineTabs.tab("Trolley Position"), placeholder_text="", width=70
        )
        self.PPStartLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Trolley Position"),
            fg_color="transparent",
            text="PP Start \n(mm)",
        )
        self.ppExtruderStartLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Trolley Position"),
            fg_color="transparent",
            text="PP Extruder \nStart (mm)",
        )
        self.ppEndDelayLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Trolley Position"),
            fg_color="transparent",
            text="PP End \nDelay (s)",
        )
        self.PPStartEntry = ctk.CTkEntry(
            self.MachineTabs.tab("Trolley Position"), placeholder_text="", width=70
        )
        self.PPEndLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Trolley Position"),
            fg_color="transparent",
            text="PP End \n(mm)",
        )
        self.PPEndEntry = ctk.CTkEntry(
            self.MachineTabs.tab("Trolley Position"), placeholder_text="", width=70
        )
        self.SocketStartLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Trolley Position"),
            fg_color="transparent",
            text="Socket Start \n(mm)",
        )
        self.SocketStartEntry = ctk.CTkEntry(
            self.MachineTabs.tab("Trolley Position"), placeholder_text="", width=70
        )
        self.SocketEndLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Trolley Position"),
            fg_color="transparent",
            text="Socket End \n(mm)",
        )
        self.SocketEndEntry = ctk.CTkEntry(
            self.MachineTabs.tab("Trolley Position"), placeholder_text="", width=70
        )
        self.pitchAngleLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Machine"),
            fg_color="transparent",
            text="Pitch Angle \n(°)",
        )
        self.pitchAngleEntry = ctk.CTkEntry(
            self.MachineTabs.tab("Machine"),
            placeholder_text="",
            width=70,
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )
        self.ppExtruderStartEntry = ctk.CTkEntry(
            self.MachineTabs.tab("Trolley Position"),
            placeholder_text="",
            width=70,
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )
        self.ppEndDelayEntry = ctk.CTkEntry(
            self.MachineTabs.tab("Trolley Position"),
            placeholder_text="",
            width=70,
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )
        self.trolleyRotataionsLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Trolley Position"),
            fg_color="transparent",
            text="Mold Rotations",
        )
        self.trolleyRotataionsEntry = ctk.CTkEntry(
            self.MachineTabs.tab("Trolley Position"), placeholder_text="", width=70
        )
        self.trolleyIncreaseTimeLabel = ctk.CTkLabel(
            self.MachineTabs.tab("Trolley Position"),
            fg_color="transparent",
            text="Trolley Increase \nTime (s)",
        )
        self.trolleyIncreaseTimeEntry = ctk.CTkEntry(
            self.MachineTabs.tab("Trolley Position"),
            placeholder_text="",
            width=70,
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )

        #

        #

        #

        #

        #

        # Results
        self.resError = ctk.CTkLabel(
            self.resultsFrame, text="", text_color="red", anchor="w"
        )
        self.moreFrame = ctk.CTkScrollableFrame(
            self.resultsFrame,
            fg_color=("#b3c7b3", "#6b756b"),
            width=600,
            height=(self.H / 4) - 20,
            orientation="horizontal",
        )
        self.moreB = ctk.CTkButton(
            self.resultsFrame,
            fg_color=("#b3c7b3", "#6b756b"),
            hover_color=("#849184", "#4f5c50"),
            text="more details",
            text_color=("black", "white"),
            command=self.moreF,
        )
        self.resultsFrame.grid_columnconfigure(4, weight=4)

        self.SnLabel = ctk.CTkLabel(
            self.moreFrame, fg_color="transparent", text="Sn:      ", anchor="w"
        )
        self.SnLabel.grid_forget()

        self.W1Label = ctk.CTkLabel(
            self.moreFrame, fg_color="transparent", text="W1:      ", anchor="w"
        )
        self.W1Label.grid_forget()

        self.W2Label = ctk.CTkLabel(
            self.moreFrame, fg_color="transparent", text="W2:      ", anchor="w"
        )
        self.W2Label.grid_forget()

        self.W3Label = ctk.CTkLabel(
            self.moreFrame, fg_color="transparent", text="W3:      ", anchor="w"
        )
        self.W3Label.grid_forget()

        self.W4Label = ctk.CTkLabel(
            self.moreFrame, fg_color="transparent", text="W4:      ", anchor="w"
        )
        self.W4Label.grid_forget()

        self.WLabel = ctk.CTkLabel(
            self.moreFrame, fg_color="transparent", text="W:      ", anchor="w"
        )
        self.WLabel.grid_forget()

        self.WkLabel = ctk.CTkLabel(
            self.moreFrame, fg_color="transparent", text="Wk:      ", anchor="w"
        )
        self.WkLabel.grid_forget()

        self.WpLabel = ctk.CTkLabel(
            self.moreFrame, fg_color="transparent", text="Wp:      ", anchor="w"
        )
        self.WpLabel.grid_forget()

        self.WzLabel = ctk.CTkLabel(
            self.moreFrame, fg_color="transparent", text="Wz:      ", anchor="w"
        )
        self.WzLabel.grid_forget()

        self.W0Label = ctk.CTkLabel(
            self.moreFrame, fg_color="transparent", text="W0:      ", anchor="w"
        )
        self.W0Label.grid_forget()

        #

        #

        # labels

        #

        #

        self.reqSnLabel = ctk.CTkLabel(
            self.productionFrame, fg_color="transparent", text="Required Sn: "
        )
        self.reqSnEntries = ctk.CTkEntry(
            self.productionFrame, placeholder_text="", width=70
        )
        self.reqSnEntries.insert(0, "8")

        # buttons

        self.optimizeButton = ctk.CTkButton(
            self.productionFrame,
            text="optimize",
            command=self.opWLs,
            width=165,
            fg_color="#5696b0",
        )

        self.calculateButton = ctk.CTkButton(
            self.productionFrame,
            text="calculate",
            command=self.calculate,
            fg_color="#5696b0",
        )
        self.productionFrame.grid_rowconfigure(5, weight=1)
        self.productionFrame.grid_columnconfigure(20, weight=1)
        self.calculateButton.grid(
            row=4, column=20, pady=5, padx=2, rowspan=2, sticky="se"
        )

        #

        #

        #
        # self.setResize(False)
        self.manyLabel = ctk.CTkLabel(
            self.sheetFrame,
            text="select Sns:",
        )

        # self.manyEntry = ctk.CTkEntry(
        #     self.sheetFrame,
        # )

        # self.reqSnLabel2 = ctk.CTkLabel(
        #     self.sheetFrame,
        #     text="Required Sns:",
        # )

        self.manyLabel.grid(row=0, column=0, padx=5)
        # self.manyEntry.grid(row=0, column=1, padx=10, sticky="nw", columnspan=2, pady=5)

        self.entrs = []

        self.expB = ctk.CTkButton(
            self.sheetFrame,
            text="Export",
            command=self.lsexp,
            fg_color="#5696b0",
        )
        self.expB.grid(
            row=1,
            column=4,
            padx=10,
            pady=5,
        )
        items = self.cur.execute("SELECT profile FROM rawMaterial").fetchall()
        items = [item[0] for item in items]
        self.matP = ctk.CTkOptionMenu(
            self.sheetFrame,
            values=list(items),
            anchor="w",
            command=self.materialCommand,
            corner_radius=0,
            dynamic_resizing=False,
            fg_color="#4eb56b",
            button_color="#2b6e3e",
            button_hover_color="#235731",
        )

        self.matP.set("Material Profiles")

        self.matP.grid(
            row=1,
            column=0,
            padx=10,
            pady=5,
            columnspan=2,
        )

        self.pFL = ctk.CTkLabel(
            self.sheetFrame,
            text="pitch factor: ",
        )
        self.pFL.grid(row=1, column=2, padx=10, pady=5, columnspan=1, sticky="e")
        self.pFE = ctk.CTkEntry(
            self.sheetFrame,
        )
        self.pFE.grid(row=1, column=3, padx=10, pady=5, columnspan=1, sticky="w")

        self.resizable(False, False)

    def lsexp(self):
        self.ls2.grid(
            row=2,
            column=0,
            columnspan=5,
            rowspan=5,
        )
        self.after(200, self.exportF)
        #

    def exportF(self):
        res = []
        self.pitchFactorEntry.configure(state="normal")
        self.pitchFactorEntry.delete(0, ctk.END)
        self.pitchFactorEntry.insert(0, str(self.pFE.get()))

        self.cur.execute("SELECT * FROM diameter")
        diameters = self.cur.fetchall()

        self.pipeLengthEnry.delete(0, ctk.END)
        self.pipeLengthEnry.insert(0, str(6000))
        tempSns = []
        for i in range(len(self.entrs)):
            t = self.entrs[i]
            if t != "" and self.isreal(t):
                tempSns.append(float(t))

        sns = {sn: [] for sn in tempSns}

        print(sns)

        Ds = {d[1]: [] for d in diameters}

        if len(tempSns) != 0:
            for d in diameters:
                self.pipeDiameterEnry.delete(0, ctk.END)
                self.pipeDiameterEnry.insert(0, str(d[1]))
                flag = False
                for sn in tempSns:
                    if flag == False:
                        Ds[d[1]].append(d[1])
                        flag = True

                    self.reqSnEntries.delete(0, ctk.END)
                    self.reqSnEntries.insert(0, str(sn))
                    self.optimizedPR()
                    fd = (
                        str(self.pitchEntry.get()[0:-2])
                        + "-S"
                        + str(self.WallThicknessEntry.get()[0:-2])
                    )
                    cd = (
                        str(self.PPDiameterEntry.get()[0:-2])
                        + "-S"
                        + str(self.PPFilmThicknessEntry.get()[0:-2])
                    )
                    if self.resError.cget("text") != "":
                        Ds[d[1]].append("-")
                        Ds[d[1]].append("-")
                        Ds[d[1]].append("-")
                    else:
                        Ds[d[1]].append(fd)
                        Ds[d[1]].append(cd)
                        Ds[d[1]].append(round(self.W0, 2))

        columns_new = []
        columns_new.append("Pipe Diameter")
        for i in range(len(tempSns)):
            columns_new.append("Flat Die")
            columns_new.append("Cladding Die")
            columns_new.append("Weight (kg/6m)")

        columns_new.append("Elastic Modulus")

        print(columns_new)

        for d in Ds.values():
            d.append(self.elastic_modulus)
        data = [d for d in Ds.values()]
        print(Ds)
        # try:
        #     writer = ExcelWriter(
        #         "Production Table " + str(self.elastic_modulus) + "MPa.xlsx", mode="a" ,if_sheet_exists="replace"
        #     )
        # except:
        #     writer = ExcelWriter(
        #         "Production Table " + str(self.elastic_modulus) + "MPa.xlsx",
        #     )

        writer = ExcelWriter(
            "Production Table " + str(self.elastic_modulus) + "MPa.xlsx",
        )

        thin = Side(border_style="thin", color="000000")

        list1 = pd.DataFrame(data, columns=columns_new)
        list1.to_excel(writer, "Optimized", index=False)
        writer.close()

        workbook = load_workbook(
            "Production Table " + str(self.elastic_modulus) + "MPa.xlsx"
        )
        ws1 = workbook["Optimized"]
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        al = Alignment(horizontal="center", vertical="center")

        ws1.insert_rows(1)

        ws1.cell(row=1, column=1).value = "Stiffness"

        for i, sn in enumerate(tempSns):
            ws1.cell(row=1, column=i * 3 + 2).value = "SN" + str(sn)
            ws1.cell(row=1, column=i * 3 + 2).font = Font(bold=True)

        ws1.cell(row=1, column=len(tempSns) * 3 + 2).value = "Remarks"
        ws1.cell(row=1, column=len(tempSns) * 3 + 2).font = Font(bold=True)
        if len(tempSns) >= 1:
            ws1.merge_cells("B1:D1")
        if len(tempSns) >= 2:
            ws1.merge_cells("E1:G1")
        if len(tempSns) >= 3:
            ws1.merge_cells("H1:J1")
        if len(tempSns) >= 4:
            ws1.merge_cells("K1:M1")
        if len(tempSns) >= 5:
            ws1.merge_cells("N1:P1")
        if len(tempSns) >= 6:
            ws1.merge_cells("Q1:S1")
            
        
   
    
        self.style_worksheet(ws1, border=border, alignment=al, itr=len(tempSns))

        workbook.save("Production Table " + str(self.elastic_modulus) + "MPa.xlsx")


        for r in res:
            print(r)
        self.pipeLengthEnry.delete(0, ctk.END)
        self.pipeDiameterEnry.delete(0, ctk.END)
        self.pitchFactorEntry.delete(0, ctk.END)
        self.reqSnEntries.delete(0, ctk.END)
        self.reqSnEntries.insert(0, "8")
        self.flatDieProfile.set("Flat die profiles")
        self.claddingDieProfile.set("Cladding die profiles")
        self.densityEntry.configure(state="normal")
        self.elasticEntry.configure(state="normal")
        self.shrinkageEntry.configure(state="normal")
        self.pitchEntry.configure(state="normal")
        self.WallThicknessEntry.configure(state="normal")
        self.finalPitchEntry.configure(state="normal")
        self.PPDiameterEntry.configure(state="normal")
        self.PPFilmThicknessEntry.configure(state="normal")
        
        self.densityEntry.delete(0, ctk.END)
        self.elasticEntry.delete(0, ctk.END)
        self.shrinkageEntry.delete(0, ctk.END)
        self.pitchEntry.delete(0, ctk.END)
        self.WallThicknessEntry.delete(0, ctk.END)
        self.finalPitchEntry.delete(0, ctk.END)
        self.PPDiameterEntry.delete(0, ctk.END)
        self.PPFilmThicknessEntry.delete(0, ctk.END)
        
        self.densityEntry.configure(state="disabled")
        self.elasticEntry.configure(state="disabled")
        self.shrinkageEntry.configure(state="disabled")
        self.pitchEntry.configure(state="disabled")
        self.WallThicknessEntry.configure(state="disabled")
        self.finalPitchEntry.configure(state="disabled")
        self.PPDiameterEntry.configure(state="disabled")
        self.PPFilmThicknessEntry.configure(state="disabled")
        for widget in self.resultsFrame.winfo_children():
            widget.grid_forget()
        for widget in self.MachineTabs.tab("Extruder").winfo_children():
            widget.grid_forget()
        for widget in self.MachineTabs.tab("Trolley Position").winfo_children():
            widget.grid_forget()
        for widget in self.MachineTabs.tab("Machine").winfo_children():
            widget.grid_forget()

        self.ls2.grid_forget()
        
        
            
        

    def style_worksheet(self, ws, border=Border(), alignment=None, itr=0):
        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)

        dims = {}

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment
                cell.border = top + left + right + bottom
                cell.font = Font(bold=False)

                
        for i in range(itr):
            ws.cell(row=1, column=i * 3 + 2).font = Font(bold=True)
        ws.cell(row=1, column=itr * 3 + 2).font = Font(bold=True)
        for i in range(itr*3+2):
            ws.cell(row=2, column=i+1).font = Font(bold=True)
        for col in ws.iter_cols(min_col = 1, max_col=1):
            for cell in col:
                cell.font = Font(bold=True)
                
        alphabet = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O", "P", "Q", "R", "S", "T"]
                
        for i, col in enumerate(ws.iter_cols()):
            lenMax = 0
            for cell in col:
                if len(str(cell.value)) > lenMax:
                    lenMax = len(str(cell.value))
                    ws.column_dimensions[alphabet[i]].width = lenMax             
        

    def buttonCommand(self):
        self.entrs = self.scrollable_checkbox_frame.get_checked_items()
        for i in range(len(self.entrs)):
            self.entrs[i] = int(self.entrs[i])

        if int(self.manyEntry.get()) > 2:
            self.expB.grid(row=2, column=2, padx=10, pady=5, sticky="nw")
            self.matP.grid(row=1, column=2, padx=10, pady=5, sticky="nw")
            self.pFE.grid(row=2, column=3, padx=10, pady=5, sticky="nw")
            self.pFL.grid(row=1, column=3, padx=10, pady=5, sticky="nw")
        else:
            self.expB.grid(
                row=2,
                column=3,
                padx=10,
                pady=5,
                columnspan=2,
            )
            self.matP.grid(
                row=1,
                column=3,
                padx=10,
                pady=5,
                columnspan=2,
            )
            self.pFE.grid(
                row=2,
                column=5,
                padx=10,
                pady=5,
                columnspan=2,
            )
            self.pFL.grid(
                row=1,
                column=5,
                padx=10,
                pady=5,
                columnspan=2,
            )

        self.reqSnLabel2.grid(row=3, column=0, sticky="nw")

    def f45(self, event):
        if self.FlatExtruder75Entry.get() != "":
            self.FlatExtruder45Entry.configure(state="normal")
            self.FlatExtruder45Entry.delete(0, ctk.END)
            self.FlatExtruder45Entry.insert(
                0, str(round(100 - float(self.FlatExtruder75Entry.get()), 2))
            )
            self.FlatExtruder45Entry.configure(
                state="disabled",
                fg_color=("#bababa", "#262626"),
            )

    def c45(self, event):
        if self.CladdingExtruder75Entry.get() != "":
            self.CladdingExtruder45Entry.configure(state="normal")
            self.CladdingExtruder45Entry.delete(0, ctk.END)
            self.CladdingExtruder45Entry.insert(
                0, str(round(100 - float(self.CladdingExtruder75Entry.get()), 2))
            )
            self.CladdingExtruder45Entry.configure(
                state="disabled",
                fg_color=("#bababa", "#262626"),
            )

    def select_frame_by_name(self, name):
        # set button color for selected button
        # self.Frame.destroy()

        self.diameterTable.configure(
            fg_color=("gray75", "gray25") if name == "diameter" else "transparent"
        )
        self.ppwtTable.configure(
            fg_color=("gray75", "gray25") if name == "ppwt" else "transparent"
        )
        self.materialProfileB.configure(
            fg_color=("gray75", "gray25") if name == "rawMaterial" else "transparent"
        )
        self.flatDiaProfileB.configure(
            fg_color=("gray75", "gray25") if name == "flatDie" else "transparent"
        )
        self.claddingDieProfileB.configure(
            fg_color=("gray75", "gray25") if name == "claddingDie" else "transparent"
        )
        self.calculatorB.configure(
            fg_color=("gray75", "gray25") if name == "calculator" else "transparent"
        )
        self.sheetB.configure(
            fg_color=("gray75", "gray25") if name == "sheet" else "transparent"
        )
        self.Frame.grid_forget()
        # show selected frame
        if name == "calculator":
            self.calcFrame.grid(row=0, column=1, sticky="nsew")
        else:
            self.calcFrame.grid_forget()
        if name == "sheet":
            self.sheetFrame.grid(row=0, column=1, sticky="nsew")
        else:
            self.sheetFrame.grid_forget()
        if name == "ppwt":
            self.refresh("ppwt", None)
            self.Frame.grid(row=0, column=1, sticky="nsew")
        elif name == "diameter":
            self.refresh("diameter", None)
            self.Frame.grid(row=0, column=1, sticky="nsew")
        elif name == "moldDiameter":
            self.refresh("moldDiameter", None)
            self.Frame.grid(row=0, column=1, sticky="nsew")
        elif name == "rawMaterial":
            self.refresh("rawMaterial", None)
            self.Frame.grid(row=0, column=1, sticky="nsew")
        elif name == "flatDie":
            self.refresh("flatDie", None)
            self.Frame.grid(row=0, column=1, sticky="nsew")
        elif name == "claddingDie":
            self.refresh("claddingDie", None)
            self.Frame.grid(row=0, column=1, sticky="nsew")

    def calcPitch(self, event):
        if self.pitchEntry.get() != "" and self.pitchFactorEntry.get() != "":
            self.finalPitchEntry.configure(state="normal")
            self.finalPitchEntry.delete(0, ctk.END)
            self.finalPitchEntry.insert(
                0,
                str(
                    round(
                        float(self.pitchEntry.get())
                        - float(self.pitchFactorEntry.get()),
                        2,
                    )
                ),
            )
            self.finalPitchEntry.configure(
                state="disabled",
                fg_color=("#bababa", "#262626"),
            )

    def Error(self, entryType, errorType, number):
        error = ctk.CTkInputDialog(
            text=(str(entryType) + " must be " + str(errorType) + ": " + str(number)),
            title="Error",
        )
        error.geometry(
            "350x200"
            + "+"
            + str(int((self.winfo_screenwidth() - 350) / 2))
            + "+"
            + str(int((self.winfo_screenheight() - 200) / 2))
        )

        flag = False

        inputNumber = error.get_input()

        if errorType.lower() == "at least":
            if float(inputNumber) >= float(number):
                flag = True
        elif errorType.lower() == "at most":
            if float(inputNumber) <= float(number):
                flag = True
        elif errorType.lower() == "equal to":
            if float(inputNumber) == float(number):
                flag = True
        elif errorType.lower() == "not equal to":
            if float(inputNumber) != float(number):
                flag = True
        else:
            error.destroy()
            return

        if flag:
            return inputNumber

        error.destroy()
        input = self.Error(entryType, errorType, number)
        return input

    def Error(self, Frame, entryName, entryType, errorType, number):
        flag = False
        if Frame == self.rawMaterialFrame:
            self.rawMaterialError.configure(
                text=(
                    str(entryType) + " must be " + str(errorType) + ": " + str(number)
                )
            )
        elif Frame == self.productionFrame:
            self.productionError.configure(
                text=(
                    str(entryType) + " must be " + str(errorType) + ": " + str(number)
                )
            )
        elif Frame == self.machineLimitsFrame:
            self.machineLimitsError.configure(
                text=(
                    str(entryType) + " must be " + str(errorType) + ": " + str(number)
                )
            )

        inputNumber = entryName.get()

        if errorType.lower() == "at least":
            if float(inputNumber) >= float(number):
                flag = True
        elif errorType.lower() == "at most":
            if float(inputNumber) <= float(number):
                flag = True
        elif errorType.lower() == "equal to":
            if float(inputNumber) == float(number):
                flag = True
        elif errorType.lower() == "not equal to":
            if float(inputNumber) != float(number):
                flag = True
        else:
            return

        if flag:
            if Frame == self.rawMaterialFrame:
                self.rawMaterialError.configure(text="")
            elif Frame == self.productionFrame:
                self.productionError.configure(text="")
            elif Frame == self.machineLimitsFrame:
                self.machineLimitsError.configure(text="")

        items = self.cur.execute("SELECT profile FROM rawMaterial").fetchall()
        items = [item[0] for item in items]
        self.materialprofile.configure(values=list(items))
        items = self.cur.execute("SELECT profile FROM flatDie").fetchall()
        items = [item[0] for item in items]
        self.flatDieProfile.configure(values=list(items))
        items = self.cur.execute("SELECT profile FROM claddingDie").fetchall()
        items = [item[0] for item in items]
        self.claddingDieProfile.configure(values=list(items))

        #

        #

        #

        #

        #

    def materialCommand(self, material):
        items = self.cur.execute("SELECT profile FROM rawMaterial").fetchall()
        items = [item[0] for item in items]
        self.materialprofile.configure(values=list(items))

        self.densityEntry.configure(state="normal")
        self.elasticEntry.configure(state="normal")
        self.shrinkageEntry.configure(state="normal")

        self.elasticEntry.delete(0, ctk.END)
        self.densityEntry.delete(0, ctk.END)
        self.shrinkageEntry.delete(0, ctk.END)
        self.elasticEntry.insert(
            0,
            str(
                self.cur.execute(
                    "SELECT elastic_modulus FROM rawMaterial WHERE profile=?",
                    (material,),
                ).fetchone()[0]
            ),
        )
        self.densityEntry.insert(
            0,
            str(
                self.cur.execute(
                    "SELECT density FROM rawMaterial WHERE profile=?", (material,)
                ).fetchone()[0]
            ),
        )
        self.shrinkageEntry.insert(
            0,
            str(
                self.cur.execute(
                    "SELECT shrinkage FROM rawMaterial WHERE profile=?", (material,)
                ).fetchone()[0]
            ),
        )
        self.densityEntry.configure(
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )
        self.elasticEntry.configure(
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )
        self.shrinkageEntry.configure(
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )

    def dieCommand(self, die):
        items = self.cur.execute("SELECT profile FROM flatDie").fetchall()
        items = [item[0] for item in items]
        self.flatDieProfile.configure(values=list(items))
        self.pitchEntry.configure(state="normal")
        self.WallThicknessEntry.configure(state="normal")
        flatdie = self.cur.execute(
            "SELECT pitch FROM flatDie WHERE profile=?", (die,)
        ).fetchone()[0]
        thickness = self.cur.execute(
            "SELECT thickness FROM flatDie WHERE profile=?", (die,)
        ).fetchone()[0]
        self.pitchEntry.delete(0, ctk.END)
        self.WallThicknessEntry.delete(0, ctk.END)
        self.WallThicknessEntry.insert(
            0,
            str(thickness),
        )
        self.pitchEntry.insert(
            0,
            str(flatdie),
        )

        self.calcPitch("")
        self.pitchEntry.configure(
            state="disabled",
        )
        self.WallThicknessEntry.configure(
            state="disabled",
        )

    # create same function as dieCommand but use ppd and ppfilmthickness ppwtEntries

    def claddingCommand(self, die):
        items = self.cur.execute("SELECT profile FROM claddingDie").fetchall()
        items = [item[0] for item in items]
        self.claddingDieProfile.configure(values=list(items))
        self.PPDiameterEntry.configure(state="normal")
        self.PPFilmThicknessEntry.configure(state="normal")
        ppd = self.cur.execute(
            "SELECT pp_diameter FROM claddingDie WHERE profile=?", (die,)
        ).fetchone()[0]
        ppfilmthickness = self.cur.execute(
            "SELECT pp_thickness FROM claddingDie WHERE profile=?", (die,)
        ).fetchone()[0]
        self.PPDiameterEntry.delete(0, ctk.END)
        self.PPFilmThicknessEntry.delete(0, ctk.END)
        self.PPDiameterEntry.insert(
            0,
            str(ppd),
        )
        self.PPFilmThicknessEntry.insert(
            0,
            str(ppfilmthickness),
        )
        self.PPDiameterEntry.configure(
            state="disabled",
        )
        self.PPFilmThicknessEntry.configure(
            state="disabled",
        )

    def moreF(self):
        if self.moreB.cget("text") == "more details":
            self.moreB.configure(text="less details")
            self.moreFrame.grid(row=0, column=4, sticky="nse", rowspan=40)
        else:
            self.moreB.configure(text="more details")
            self.moreFrame.grid_forget()

    def modeCommand(self, mode):
        self.mode = mode
        # clear all widgets in machine frame and results frame
        for widget in self.resultsFrame.winfo_children():
            widget.grid_forget()
        for widget in self.MachineTabs.tab("Extruder").winfo_children():
            widget.grid_forget()
        for widget in self.MachineTabs.tab("Trolley Position").winfo_children():
            widget.grid_forget()
        for widget in self.MachineTabs.tab("Machine").winfo_children():
            widget.grid_forget()
        self.materialprofile.grid_forget()
        self.densityLabel.grid_forget()
        self.densityEntry.grid_forget()
        self.densityUnitLabel.grid_forget()
        self.elasticLabel.grid_forget()
        self.elasticEntry.grid_forget()
        self.elasticUnitLabel.grid_forget()
        self.shrinkageLabel.grid_forget()
        self.shrinkageEntry.grid_forget()
        self.shrinkageUnitLabel.grid_forget()

        self.flatDieProfile.grid_forget()
        self.claddingDieProfile.grid_forget()
        self.optimizeButton.grid_forget()
        self.reqSnLabel.grid_forget()
        self.reqSnEntries.grid_forget()
        self.pipeLengthLabel.grid_forget()
        self.pipeLengthUnitLabel.grid_forget()
        self.pipeLengthEnry.grid_forget()
        self.pipeDiameterLabel.grid_forget()
        self.pipeDiameterUnitLabel.grid_forget()
        self.pipeDiameterEnry.grid_forget()
        self.pitchLabel.grid_forget()
        self.pitchEntry.grid_forget()
        self.pitchUnitLabel.grid_forget()
        self.pitchFactorLabel.grid_forget()
        self.pitchFactorEntry.grid_forget()
        self.pitchFactorUnitLabel.grid_forget()
        self.WallThicknessLabel.grid_forget()
        self.WallThicknessEntry.grid_forget()
        self.WallThicknessUnitLabel.grid_forget()
        self.finalPitchLabel.grid_forget()
        self.finalPitchEntry.grid_forget()
        self.finalPitchUnitLabel.grid_forget()
        self.PPDiameterLabel.grid_forget()
        self.PPDiameterEntry.grid_forget()
        self.PPDiameterUnitLabel.grid_forget()
        self.PPFilmThicknessLabel.grid_forget()
        self.PPFilmThicknessEntry.grid_forget()
        self.PPFilmThicknessUnitLabel.grid_forget()

        self.resError.grid_forget()
        self.SnL.grid_forget()
        self.W0L.grid_forget()
        self.WC.grid_forget()
        self.WP.grid_forget()
        self.moreB.grid_forget()
        self.moreB.configure(text="more details")
        self.moreFrame.grid_forget()

        self.maxFactorLabel.grid_forget()
        self.maxFactorUnitLabel.grid_forget()
        self.maxFactorEntry.grid_forget()
        self.flat75MaxLabel.grid_forget()
        self.flat75MaxUnitLabel.grid_forget()
        self.flat75MaxEntry.grid_forget()
        self.flat45MaxLabel.grid_forget()
        self.flat45MaxUnitLabel.grid_forget()
        self.flat45MaxEntry.grid_forget()
        self.cladding75MaxLabel.grid_forget()
        self.cladding75MaxUnitLabel.grid_forget()
        self.cladding75MaxEntry.grid_forget()
        self.cladding45MaxLabel.grid_forget()
        self.cladding45MaxUnitLabel.grid_forget()
        self.cladding45MaxEntry.grid_forget()
        self.MoldSpeedLabel.grid_forget()
        self.MoldSpeedUnitLabel.grid_forget()
        self.MoldSpeedEntry.grid_forget()
        self.ppSpeedFractionLabel.grid_forget()
        self.ppSpeedFractionUnitLabel.grid_forget()
        self.ppSpeedFractionEntry.grid_forget()
        self.ppSpeedLabel.grid_forget()
        self.ppSpeedUnitLabel.grid_forget()
        self.ppSpeedEntry.grid_forget()
        self.trolleySpeedLabel.grid_forget()
        self.trolleySpeedUnitLabel.grid_forget()
        self.trolleySpeedEntry.grid_forget()
        self.FlatExtruder75Label.grid_forget()
        self.FlatExtruder75UnitLabel.grid_forget()
        self.FlatExtruder75Entry.grid_forget()
        self.FlatExtruder45Label.grid_forget()
        self.FlatExtruder45UnitLabel.grid_forget()
        self.FlatExtruder45Entry.grid_forget()
        self.CladdingExtruder75Label.grid_forget()
        self.CladdingExtruder75UnitLabel.grid_forget()
        self.CladdingExtruder75Entry.grid_forget()
        self.CladdingExtruder45Label.grid_forget()
        self.CladdingExtruder45UnitLabel.grid_forget()
        self.CladdingExtruder45Entry.grid_forget()

        if mode == "PR":
            self.PR()
        elif mode == "VW":
            self.VW()

        #

        #

        #

        #

        #

    def change_appearance_mode_event(self, new_appearance_mode):
        ctk.set_appearance_mode(new_appearance_mode)

    def PR(self):  # show widgets related to PR
        self.materialprofile.grid(row=1, column=2, columnspan=2, sticky="w", padx=5)
        self.flatDieProfile.grid(row=2, column=0, columnspan=2, sticky="w", padx=5)
        self.claddingDieProfile.grid(row=3, column=0, columnspan=2, sticky="w", padx=5)

        self.optimizeButton.grid(
            row=1, column=0, columnspan=2, sticky="w", padx=5, pady=5
        )
        self.reqSnLabel.grid(row=1, column=16, columnspan=1, sticky="w", padx=5)
        self.reqSnEntries.grid(row=1, column=17, columnspan=1, padx=5)

        self.densityLabel.grid(row=1, column=5, padx=5, pady=5)
        self.densityEntry.grid(row=1, column=6, padx=5, pady=5)
        self.densityUnitLabel.grid(row=1, column=7, padx=1, pady=5, sticky="w")
        self.elasticLabel.grid(row=1, column=8, padx=20, pady=5, columnspan=2)
        self.elasticEntry.grid(row=1, column=10, padx=2, pady=5, sticky="w")
        self.elasticUnitLabel.grid(row=1, column=11, padx=1, pady=5, sticky="w")
        self.shrinkageLabel.grid(row=1, column=12, padx=20, pady=5, columnspan=2)
        self.shrinkageEntry.grid(row=1, column=14, padx=2, pady=5, sticky="w")
        self.shrinkageUnitLabel.grid(row=1, column=15, padx=1, pady=5, sticky="w")

        self.pipeLengthLabel.grid(row=1, column=2, padx=5, pady=5, sticky="w")
        self.pipeLengthUnitLabel.grid(row=1, column=4, padx=1, pady=5, sticky="w")
        self.pipeLengthEnry.grid(row=1, column=3, padx=6, pady=5)
        self.pipeDiameterLabel.grid(row=1, column=5, padx=10, pady=5, sticky="w")
        self.pipeDiameterUnitLabel.grid(row=1, column=7, padx=1, pady=5, sticky="w")
        self.pipeDiameterEnry.grid(row=1, column=6, padx=6, pady=5)
        self.pitchFactorLabel.grid(row=1, column=8, padx=5, pady=5, sticky="w")
        self.pitchFactorEntry.grid(row=1, column=9, padx=6, pady=5)
        self.pitchFactorUnitLabel.grid(row=1, column=10, padx=1, pady=5, sticky="w")
        self.pitchLabel.grid(row=2, column=2, padx=5, pady=5, sticky="w")
        self.pitchEntry.grid(row=2, column=3, padx=6, pady=5)
        self.pitchUnitLabel.grid(row=2, column=4, padx=1, pady=5, sticky="w")
        self.WallThicknessLabel.grid(row=2, column=5, padx=10, pady=5, sticky="w")
        self.WallThicknessEntry.grid(row=2, column=6, padx=6, pady=5)
        self.WallThicknessUnitLabel.grid(row=2, column=7, padx=1, pady=5, sticky="w")
        self.finalPitchLabel.grid(row=2, column=8, padx=5, pady=5, sticky="w")
        self.finalPitchEntry.grid(row=2, column=9, padx=6, pady=5)
        self.finalPitchUnitLabel.grid(row=2, column=10, padx=1, pady=5, sticky="w")
        self.PPDiameterLabel.grid(row=3, column=2, padx=5, pady=5, sticky="w")
        self.PPDiameterEntry.grid(row=3, column=3, padx=6, pady=5)
        self.PPDiameterUnitLabel.grid(row=3, column=4, padx=1, pady=5, sticky="w")
        self.PPFilmThicknessLabel.grid(row=3, column=5, padx=10, pady=5, sticky="w")
        self.PPFilmThicknessEntry.grid(row=3, column=6, padx=6, pady=5)
        self.PPFilmThicknessUnitLabel.grid(row=3, column=7, padx=1, pady=5, sticky="w")

        #

        #

        #

        #

        #

    def VW(self):
        self.materialprofile.grid(row=1, column=2, columnspan=2, sticky="w", padx=5)
        self.flatDieProfile.grid(row=2, column=0, columnspan=2, sticky="w", padx=5)

        self.densityLabel.grid(row=1, column=5, padx=5, pady=5)
        self.densityEntry.grid(row=1, column=6, padx=5, pady=5)
        self.densityUnitLabel.grid(row=1, column=7, padx=1, pady=5, sticky="w")
        self.elasticLabel.grid(row=1, column=8, padx=20, pady=5, columnspan=2)
        self.elasticEntry.grid(row=1, column=10, padx=2, pady=5, sticky="w")
        self.elasticUnitLabel.grid(row=1, column=11, padx=1, pady=5, sticky="w")
        self.shrinkageLabel.grid(row=1, column=12, padx=20, pady=5, columnspan=2)
        self.shrinkageEntry.grid(row=1, column=14, padx=2, pady=5, sticky="w")
        self.shrinkageUnitLabel.grid(row=1, column=15, padx=1, pady=5, sticky="w")

        self.pipeLengthLabel.grid(row=1, column=2, padx=5, pady=5, sticky="w")
        self.pipeLengthUnitLabel.grid(row=1, column=4, padx=1, pady=5, sticky="w")
        self.pipeLengthEnry.grid(row=1, column=3, padx=6, pady=5)
        self.pipeDiameterLabel.grid(row=1, column=5, padx=10, pady=5, sticky="w")
        self.pipeDiameterUnitLabel.grid(row=1, column=7, padx=1, pady=5, sticky="w")
        self.pipeDiameterEnry.grid(row=1, column=6, padx=6, pady=5)
        self.pitchFactorLabel.grid(row=1, column=8, padx=5, pady=5, sticky="w")
        self.pitchFactorEntry.grid(row=1, column=9, padx=6, pady=5)
        self.pitchFactorUnitLabel.grid(row=1, column=10, padx=1, pady=5, sticky="w")
        self.pitchLabel.grid(row=2, column=2, padx=5, pady=5, sticky="w")
        self.pitchEntry.grid(row=2, column=3, padx=6, pady=5)
        self.pitchUnitLabel.grid(row=2, column=4, padx=1, pady=5, sticky="w")
        self.WallThicknessLabel.grid(row=2, column=5, padx=10, pady=5, sticky="w")
        self.WallThicknessEntry.grid(row=2, column=6, padx=6, pady=5)
        self.WallThicknessUnitLabel.grid(row=2, column=7, padx=1, pady=5, sticky="w")
        self.finalPitchLabel.grid(row=2, column=8, padx=5, pady=5, sticky="w")
        self.finalPitchEntry.grid(row=2, column=9, padx=6, pady=5)
        self.finalPitchUnitLabel.grid(row=2, column=10, padx=1, pady=5, sticky="w")

    def calcSpeeds(self, event=None):
        self.FlatExtruder75EntryF.configure(state="normal")
        self.FlatExtruder75EntryF.delete(0, ctk.END)
        self.FlatExtruder75EntryF.configure(
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )
        self.FlatExtruder45EntryF.configure(state="normal")
        self.FlatExtruder45EntryF.delete(0, ctk.END)
        self.FlatExtruder45EntryF.configure(
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )
        self.CladdingExtruder75EntryF.configure(state="normal")
        self.CladdingExtruder75EntryF.delete(0, ctk.END)
        self.CladdingExtruder75EntryF.configure(
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )
        self.CladdingExtruder45EntryF.configure(state="normal")
        self.CladdingExtruder45EntryF.delete(0, ctk.END)
        self.CladdingExtruder45EntryF.configure(
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )
        try:
            temP = self.MoldSpeedEntry.get()
            if self.isreal(temP) == False:
                raise Exception("Mold speed must be a number")
            temp = (self.pd / float(temP)) * 60 * math.pi
            if self.isreal(self.CarriageRotationsEntry.get()) == False:
                raise Exception("Carriage rotations must be a number")
            if self.isreal(self.trolleyRotataionsEntry.get()) == False:
                raise Exception("Trolley rotations must be a number")
            self.CarriageReturnDelayEntry.configure(state="normal")
            self.CarriageReturnDelayEntry.delete(0, ctk.END)
            self.CarriageReturnDelayEntry.insert(
                0, str(round(temp * float(self.CarriageRotationsEntry.get()), 2))
            )
            self.CarriageReturnDelayEntry.configure(
                state="disabled",
                fg_color=("#bababa", "#262626"),
            )
            self.trolleyIncreaseTimeEntry.configure(state="normal")
            self.trolleyIncreaseTimeEntry.delete(0, ctk.END)
            self.trolleyIncreaseTimeEntry.insert(
                0, str(round(temp * float(self.trolleyRotataionsEntry.get()), 2))
            )
            self.trolleyIncreaseTimeEntry.configure(
                state="disabled",
                fg_color=("#bababa", "#262626"),
            )
            self.calcTrolley(None)
            try:
                self.trolleyError.destroy()
            except:
                pass
        except Exception as e:
            self.trolleyError = ctk.CTkLabel(
                self.MachineTabs.tab("Trolley Position"),
                text_color="red",
                text=str(e),
            )
            self.trolleyError.grid(
                row=3, column=0, padx=10, pady=5, columnspan=10, sticky="w"
            )

        self.machineLimitsError.configure(text="")
        self.ppSpeedEntry.configure(state="normal")
        self.ppSpeedEntry.delete(0, ctk.END)
        temp = float(self.MoldSpeedEntry.get()) * (
            float(self.ppSpeedFractionEntry.get()) / 100.0
        )
        self.ppSpeedEntry.insert(0, str(round(temp, 2)))
        self.ppSpeedEntry.configure(
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )
        self.trolleySpeedEntry.configure(state="normal")
        self.trolleySpeedEntry.delete(0, ctk.END)
        temp = (float(self.MoldSpeedEntry.get()) * float(self.p)) / (
            float(self.pd) * math.pi
        )
        self.trolleySpeedEntry.insert(0, str(round(temp, 2)))
        self.trolleySpeedEntry.configure(
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )

        flatExtruder = (
            60
            / float(self.pipeLengthEnry.get())
            * float(self.trolleySpeedEntry.get())
            * self.W3
        )
        claddingExtruder = (
            60
            / float(self.pipeLengthEnry.get())
            * float(self.trolleySpeedEntry.get())
            * self.W4
        )
        try:
            self.FlatExtruder75Entry.configure(
                fg_color=("white", "grey20"), border_color=("#999EA3", "grey30")
            )
            self.FlatExtruder45Entry.configure(
                fg_color=("#bababa", "#262626"), border_color=("#999EA3", "grey30")
            )
            self.cladding75MaxEntry.configure(
                fg_color=("white", "grey20"), border_color=("#999EA3", "grey30")
            )
            self.cladding45MaxEntry.configure(
                fg_color=("white", "grey20"), border_color=("#999EA3", "grey30")
            )
            # check if inputs are numbers
            if self.isreal(self.maxFactorEntry.get()) == False:
                self.maxFactorEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Max factor must be a number")
            if self.isreal(self.flat75MaxEntry.get()) == False:
                self.flat75MaxEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Flat extruder 75 max must be a number")
            if self.isreal(self.flat45MaxEntry.get()) == False:
                self.flat45MaxEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Flat extruder 45 max must be a number")
            if self.isreal(self.cladding75MaxEntry.get()) == False:
                self.cladding75MaxEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Cladding extruder 75 max must be a number")
            if self.isreal(self.cladding45MaxEntry.get()) == False:
                self.cladding45MaxEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Cladding extruder 45 max must be a number")
            if self.isreal(self.MoldSpeedEntry.get()) == False:
                self.MoldSpeedEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Mold speed must be a number")
            if self.isreal(self.ppSpeedFractionEntry.get()) == False:
                self.ppSpeedFractionEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("PP speed fraction must be a number")
            if self.isreal(self.trolleySpeedEntry.get()) == False:
                self.trolleySpeedEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Trolley speed must be a number")
            if self.isreal(self.trolleyIncreaseTimeEntry.get()) == False:
                self.trolleyIncreaseTimeEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Trolley increase time must be a number")
            if self.isreal(self.FlatExtruder75Entry.get()) == False:
                self.FlatExtruder75Entry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Flat extruder 75 must be a number")
            if self.isreal(self.FlatExtruder45Entry.get()) == False:
                self.FlatExtruder45Entry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Flat extruder 45 must be a number")
            if self.isreal(self.CladdingExtruder75Entry.get()) == False:
                self.CladdingExtruder75Entry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Cladding extruder 75 must be a number")
            if self.isreal(self.CladdingExtruder45Entry.get()) == False:
                self.CladdingExtruder45Entry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Cladding extruder 45 must be a number")

            flatExtruder75 = flatExtruder * float(self.FlatExtruder75Entry.get()) / 100
            self.FlatExtruder75EntryF.configure(state="normal")
            self.FlatExtruder75EntryF.delete(0, ctk.END)
            self.FlatExtruder75EntryF.insert(0, str(round(flatExtruder75, 2)))
            self.FlatExtruder75EntryF.configure(
                state="disabled",
                fg_color=("#bababa", "#262626"),
            )
            flatExtruder45 = flatExtruder * float(self.FlatExtruder45Entry.get()) / 100
            self.FlatExtruder45EntryF.configure(state="normal")
            self.FlatExtruder45EntryF.delete(0, ctk.END)
            self.FlatExtruder45EntryF.insert(0, str(round(flatExtruder45, 2)))
            self.FlatExtruder45EntryF.configure(
                state="disabled",
                fg_color=("#bababa", "#262626"),
            )
            claddingExtruder75 = (
                claddingExtruder * float(self.CladdingExtruder75Entry.get()) / 100
            )
            self.CladdingExtruder75EntryF.configure(state="normal")
            self.CladdingExtruder75EntryF.delete(0, ctk.END)
            self.CladdingExtruder75EntryF.insert(0, str(round(claddingExtruder75, 2)))
            self.CladdingExtruder75EntryF.configure(
                state="disabled",
                fg_color=("#bababa", "#262626"),
            )
            claddingExtruder45 = (
                claddingExtruder * float(self.CladdingExtruder45Entry.get()) / 100
            )
            self.CladdingExtruder45EntryF.configure(state="normal")
            self.CladdingExtruder45EntryF.delete(0, ctk.END)
            self.CladdingExtruder45EntryF.insert(0, str(round(claddingExtruder45, 2)))
            self.CladdingExtruder45EntryF.configure(
                state="disabled",
                fg_color=("#bababa", "#262626"),
            )

            self.FlatExtruder75Entry.bind("<Return>", self.calcSpeeds)

            self.flat75MaxEntry.bind("<Return>", self.calcSpeeds)

            self.FlatExtruder45Entry.bind("<Return>", self.calcSpeeds)

            self.flat45MaxEntry.bind("<Return>", self.calcSpeeds)

            self.CladdingExtruder75Entry.bind("<Return>", self.calcSpeeds)

            self.cladding75MaxEntry.bind("<Return>", self.calcSpeeds)

            self.CladdingExtruder45Entry.bind("<Return>", self.calcSpeeds)

            self.cladding45MaxEntry.bind("<Return>", self.calcSpeeds)
            if (
                flatExtruder75
                > float(self.flat75MaxEntry.get())
                * float(self.maxFactorEntry.get())
                / 100
            ):
                self.FlatExtruder75Entry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                self.flat75MaxEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Flat extruder 75 exceeded max")
            self.FlatExtruder75Entry.configure(
                fg_color=("white", "grey20"),
                border_color=("#999EA3", "grey30"),
            )
            self.flat75MaxEntry.configure(
                fg_color=("white", "grey20"),
                border_color=("#999EA3", "grey30"),
            )

            if (
                flatExtruder45
                > float(self.flat45MaxEntry.get())
                * float(self.maxFactorEntry.get())
                / 100
            ):
                self.FlatExtruder45Entry.configure(
                    fg_color=("#bababa", "#262626"),
                    border_color=("#999EA3", "grey30"),
                )
                self.flat45MaxEntry.configure(
                    fg_color=("white", "grey20"),
                    border_color=("#999EA3", "grey30"),
                )

                raise Exception("Flat extruder 45 exceeded max")
            self.FlatExtruder45Entry.configure(
                fg_color=("#bababa", "#262626"),
                border_color=("#999EA3", "grey30"),
            )
            self.flat45MaxEntry.configure(
                fg_color=("white", "grey20"),
                border_color=("#999EA3", "grey30"),
            )

            if (
                claddingExtruder75
                > float(self.cladding75MaxEntry.get())
                * float(self.maxFactorEntry.get())
                / 100
            ):
                self.CladdingExtruder75Entry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                self.cladding75MaxEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )

                raise Exception("Clad extruder 75 exceeded max")
            self.CladdingExtruder75Entry.configure(
                fg_color=("white", "grey20"),
                border_color=("#999EA3", "grey30"),
            )
            self.cladding75MaxEntry.configure(
                fg_color=("white", "grey20"),
                border_color=("#999EA3", "grey30"),
            )

            if (
                claddingExtruder45
                > float(self.cladding45MaxEntry.get())
                * float(self.maxFactorEntry.get())
                / 100
            ):
                self.CladdingExtruder45Entry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                self.cladding45MaxEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )

                raise Exception("Clad extruder 45 exceeded max")
            self.CladdingExtruder45Entry.configure(
                fg_color=("#bababa", "#262626"), border_color=("#999EA3", "grey30")
            )
            self.cladding45MaxEntry.configure(
                fg_color=("white", "grey20"), border_color=("#999EA3", "grey30")
            )
        except Exception as e:
            self.machineLimitsError.configure(text=str(e))

    def isreal(self, s):
        try:
            float(s)
            return True
        except ValueError:
            return False

    def CS(self, entry, event):
        entry.bind("<FocusOut", self.unb)

    def calculate(self):
        self.SnLabel.configure(text="")
        self.W1Label.configure(text="")
        self.W2Label.configure(text="")
        self.W3Label.configure(text="")
        self.W4Label.configure(text="")
        self.WLabel.configure(text="")
        self.WkLabel.configure(text="")
        self.WpLabel.configure(text="")
        self.WzLabel.configure(text="")
        self.W0Label.configure(text="")
        self.SnL.configure(text="")
        self.W0L.configure(text="")
        self.WC.configure(text="")
        self.WP.configure(text="")
        self.resError.configure(text="")
        self.resError.grid_forget()
        self.moreB.grid_forget()
        self.moreB.configure(text="more details")

        try:
            if self.isreal(self.densityEntry.get()) == False:
                self.densityEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Density must be a number")
            if float(self.densityEntry.get()) < 0:
                self.densityEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Density must be positive")
            self.densityEntry.configure(
                fg_color=("#bababa", "#262626"), border_color=("#999EA3", "grey30")
            )
            if self.isreal(self.elasticEntry.get()) == False:
                self.elasticEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Elastic Modulus must be a number")
            if float(self.elasticEntry.get()) < 0:
                self.elasticEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Elastic Modulus must be positive")
            self.elasticEntry.configure(
                fg_color=("#bababa", "#262626"), border_color=("#999EA3", "grey30")
            )
            if self.isreal(self.shrinkageEntry.get()) == False:
                self.shrinkageEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Shrinkage must be a number")
            if float(self.shrinkageEntry.get()) < 0:
                self.shrinkageEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Shrinkage must be positive")
            self.shrinkageEntry.configure(
                fg_color=("#bababa", "#262626"), border_color=("#999EA3", "grey30")
            )
            if self.isreal(self.pipeLengthEnry.get()) == False:
                self.pipeLengthEnry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Pipe Lenght must be a number")
            if float(self.pipeLengthEnry.get()) < 0:
                self.pipeLengthEnry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Pipe Lenght must be positive")
            self.pipeLengthEnry.configure(
                fg_color=("white", "grey20"), border_color=("#999EA3", "grey30")
            )
            if self.isreal(self.pipeDiameterEnry.get()) == False:
                self.pipeDiameterEnry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Pipe Diameter must be a number")
            if float(self.pipeDiameterEnry.get()) < 0:
                self.pipeDiameterEnry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Pipe Diameter must be positive")
            self.pipeDiameterEnry.configure(
                fg_color=("white", "grey20"), border_color=("#999EA3", "grey30")
            )
            if self.isreal(self.pitchFactorEntry.get()) == False:
                self.pitchFactorEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Pitch Factor must be a number")
            if float(self.pitchFactorEntry.get()) < 0:
                self.pitchFactorEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Pitch Factor must be positive")
            self.pitchFactorEntry.configure(
                fg_color=("white", "grey20"), border_color=("#999EA3", "grey30")
            )
            if self.isreal(self.pitchEntry.get()) == False:
                self.pitchEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Pitch must be a number")
            if float(self.pitchEntry.get()) < 0:
                self.pitchEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Pitch must be positive")
            self.pitchEntry.configure(
                fg_color=("#bababa", "#262626"), border_color=("#999EA3", "grey30")
            )
        except Exception as e:
            self.productionError.configure(text=str(e))
            return

        if self.mode == "PR":
            self.calculatePR()
        elif self.mode == "VW":
            self.calculateVW()

        self.maxFactorLabel.grid(row=0, column=1, padx=10, pady=5, sticky="w")
        self.maxFactorUnitLabel.grid(row=0, column=3, padx=10, pady=5)
        self.maxFactorEntry.grid(row=0, column=2, padx=10, pady=5, sticky="w")
        self.flat75MaxLabel.grid(row=1, column=1, padx=10, pady=5, sticky="w")
        self.flat75MaxUnitLabel.grid(row=1, column=3, padx=10, pady=5)
        self.flat75MaxEntry.grid(row=1, column=2, padx=10, pady=5, sticky="w")
        self.flat45MaxLabel.grid(row=2, column=1, padx=10, pady=5, sticky="w")
        self.flat45MaxUnitLabel.grid(row=2, column=3, padx=10, pady=5)
        self.flat45MaxEntry.grid(row=2, column=2, padx=10, pady=5, sticky="w")
        self.cladding75MaxLabel.grid(row=3, column=1, padx=10, pady=5, sticky="w")
        self.cladding75MaxUnitLabel.grid(row=3, column=3, padx=10, pady=5)
        self.cladding75MaxEntry.grid(row=3, column=2, padx=10, pady=5, sticky="w")
        self.cladding45MaxLabel.grid(row=4, column=1, padx=10, pady=5, sticky="w")
        self.cladding45MaxUnitLabel.grid(row=4, column=3, padx=10, pady=5)
        self.cladding45MaxEntry.grid(row=4, column=2, padx=10, pady=5, sticky="w")
        self.FlatExtruder75Label.grid(row=1, column=4, padx=10, pady=5, sticky="w")
        self.FlatExtruder75UnitLabel.grid(row=1, column=6, padx=10, pady=5)
        self.FlatExtruder75Entry.grid(row=1, column=5, padx=10, pady=5, sticky="w")
        self.FlatExtruder45Label.grid(row=2, column=4, padx=10, pady=5, sticky="w")
        self.FlatExtruder45UnitLabel.grid(row=2, column=6, padx=10, pady=5)
        self.FlatExtruder45Entry.grid(row=2, column=5, padx=10, pady=5, sticky="w")
        self.CladdingExtruder75Label.grid(row=3, column=4, padx=10, pady=5, sticky="w")
        self.CladdingExtruder75UnitLabel.grid(row=3, column=6, padx=10, pady=5)
        self.CladdingExtruder75Entry.grid(row=3, column=5, padx=10, pady=5, sticky="w")
        self.CladdingExtruder45Label.grid(row=4, column=4, padx=10, pady=5, sticky="w")
        self.CladdingExtruder45UnitLabel.grid(row=4, column=6, padx=10, pady=5)
        self.CladdingExtruder45Entry.grid(row=4, column=5, padx=10, pady=5, sticky="w")

        self.FlatExtruder75LabelF.grid(row=1, column=7, padx=10, pady=5, sticky="w")
        self.FlatExtruder75UnitLabelF.grid(row=1, column=9, padx=10, pady=5)
        self.FlatExtruder75EntryF.grid(row=1, column=8, padx=10, pady=5, sticky="w")
        self.FlatExtruder45LabelF.grid(row=2, column=7, padx=10, pady=5, sticky="w")
        self.FlatExtruder45UnitLabelF.grid(row=2, column=9, padx=10, pady=5)
        self.FlatExtruder45EntryF.grid(row=2, column=8, padx=10, pady=5, sticky="w")
        self.CladdingExtruder75LabelF.grid(row=3, column=7, padx=10, pady=5, sticky="w")
        self.CladdingExtruder75UnitLabelF.grid(row=3, column=9, padx=10, pady=5)
        self.CladdingExtruder75EntryF.grid(row=3, column=8, padx=10, pady=5, sticky="w")
        self.CladdingExtruder45LabelF.grid(row=4, column=7, padx=10, pady=5, sticky="w")
        self.CladdingExtruder45UnitLabelF.grid(row=4, column=9, padx=10, pady=5)
        self.CladdingExtruder45EntryF.grid(row=4, column=8, padx=10, pady=5, sticky="w")

        self.MoldSpeedLabel.grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.MoldSpeedUnitLabel.grid(row=0, column=2, padx=10, pady=5)
        self.MoldSpeedEntry.grid(row=0, column=1, padx=10, pady=5, sticky="w")
        self.ppSpeedLabel.grid(row=1, column=3, padx=10, pady=5, sticky="w")
        self.ppSpeedUnitLabel.grid(row=1, column=5, padx=10, pady=5)
        self.ppSpeedEntry.grid(row=1, column=4, padx=10, pady=5, sticky="w")
        self.trolleySpeedLabel.grid(row=0, column=3, padx=10, pady=5, sticky="w")
        self.trolleySpeedUnitLabel.grid(row=0, column=5, padx=10, pady=5)
        self.trolleySpeedEntry.grid(row=0, column=4, padx=10, pady=5, sticky="w")
        self.ppSpeedFractionLabel.grid(row=1, column=0, padx=10, pady=5, sticky="w")
        self.ppSpeedFractionUnitLabel.grid(row=1, column=2, padx=10, pady=5)
        self.ppSpeedFractionEntry.grid(row=1, column=1, padx=10, pady=5, sticky="w")

        self.MouthStartLabel.grid(
            row=0,
            column=0,
            padx=10,
            pady=5,
        )
        self.MouthEndLabel.grid(
            row=0,
            column=2,
            padx=10,
            pady=5,
        )
        self.CarriageReturnPositionLabel.grid(
            row=0,
            column=4,
            padx=10,
            pady=5,
        )
        self.CarriageRotationsLabel.grid(
            row=0,
            column=6,
            padx=10,
            pady=5,
        )
        self.CarriageReturnDelayLabel.grid(
            row=0,
            column=8,
            padx=10,
            pady=5,
        )
        self.PPStartLabel.grid(
            row=1,
            column=0,
            padx=10,
            pady=5,
        )
        self.PPEndLabel.grid(
            row=1,
            column=2,
            padx=10,
            pady=5,
        )
        self.SocketStartLabel.grid(
            row=1,
            column=4,
            padx=10,
            pady=5,
        )
        self.SocketEndLabel.grid(
            row=1,
            column=6,
            padx=10,
            pady=5,
        )
        self.pitchAngleLabel.grid(
            row=2,
            column=0,
            padx=10,
            pady=5,
        )
        self.trolleyRotataionsLabel.grid(
            row=2,
            column=0,
            padx=10,
            pady=5,
        )
        self.trolleyIncreaseTimeLabel.grid(
            row=2,
            column=2,
            padx=10,
            pady=5,
        )
        self.ppExtruderStartLabel.grid(
            row=2,
            column=4,
            padx=10,
            pady=5,
        )
        self.ppEndDelayLabel.grid(
            row=2,
            column=6,
            padx=10,
            pady=5,
        )

        self.MouthStartEntry.grid(
            row=0,
            column=1,
            padx=10,
            pady=5,
        )
        self.MouthStartEntry.delete(0, ctk.END)
        self.MouthStartEntry.insert(0, "115")
        self.MouthEndEntry.grid(row=0, column=3, padx=10, pady=5)
        self.MouthEndEntry.delete(0, ctk.END)
        self.MouthEndEntry.insert(0, "680")
        self.CarriageReturnPositionEntry.delete(0, ctk.END)
        self.CarriageReturnPositionEntry.insert(0, "350")
        self.CarriageReturnPositionEntry.grid(
            row=0,
            column=5,
            padx=10,
            pady=5,
        )
        self.CarriageRotationsEntry.delete(0, ctk.END)
        self.CarriageRotationsEntry.insert(0, "1")
        self.CarriageRotationsEntry.grid(
            row=0,
            column=7,
            padx=10,
            pady=5,
        )

        self.CarriageRotationsEntry.bind("<Return>", self.calcSpeeds)
        self.CarriageReturnDelayEntry.grid(
            row=0,
            column=9,
            padx=10,
            pady=5,
        )
        self.PPStartEntry.grid(
            row=1,
            column=1,
            padx=10,
            pady=5,
        )

        self.PPStartEntry.bind("<Return>", self.calcPPExSt)
        self.PPStartEntry.delete(0, ctk.END)
        self.PPStartEntry.insert(0, "680")
        self.PPEndEntry.grid(
            row=1,
            column=3,
            padx=10,
            pady=5,
        )
        self.PPEndEntry.delete(0, ctk.END)
        self.PPEndEntry.insert(0, "6114")
        self.SocketStartEntry.grid(
            row=1,
            column=5,
            padx=10,
            pady=5,
        )
        self.SocketStartEntry.delete(0, ctk.END)
        self.SocketStartEntry.insert(0, "6000")
        self.SocketEndEntry.grid(row=1, column=7, padx=10, pady=5)
        self.SocketEndEntry.delete(0, ctk.END)
        self.SocketEndEntry.insert(0, "6115")
        temp = math.degrees(
            math.atan((self.p / 4) / (self.pd / 2 + self.s1 + self.ppd + self.s4 * 2))
        )
        self.pitchAngleEntry.configure(state="normal")
        self.pitchAngleEntry.delete(0, ctk.END)
        self.pitchAngleEntry.insert(0, str(round(temp, 2)))
        self.pitchAngleEntry.grid(row=2, column=1, padx=10, pady=5)
        self.pitchAngleEntry.configure(state="disabled")
        self.trolleyRotataionsEntry.grid(
            row=2,
            column=1,
            padx=10,
            pady=5,
        )
        self.trolleyRotataionsEntry.delete(0, ctk.END)
        self.trolleyRotataionsEntry.insert(0, "1")

        self.trolleyRotataionsEntry.bind("<Return>", self.calcSpeeds)
        self.trolleyIncreaseTimeEntry.grid(
            row=2,
            column=3,
            padx=10,
            pady=5,
        )

        self.trolleyIncreaseTimeEntry.bind("<Return>", self.calcTrolley)
        self.ppExtruderStartEntry.grid(
            row=2,
            column=5,
            padx=10,
            pady=5,
        )
        self.ppEndDelayEntry.grid(
            row=2,
            column=7,
            padx=10,
            pady=5,
        )
        self.f45(None)
        self.c45(None)
        self.calcSpeeds(None)
        self.calcPPExSt(None)

    def calcPPExSt(self, event=None):
        try:
            self.ppExtruderStartEntry.configure(state="normal")
            self.ppExtruderStartEntry.delete(0, ctk.END)
            self.ppExtruderStartEntry.insert(
                0, str(float(self.PPStartEntry.get()) - 200)
            )
            self.ppExtruderStartEntry.configure(
                state="disabled",
                fg_color=("#bababa", "#262626"),
            )
            self.PPStartEntry.configure(
                fg_color=("white", "grey20"), border_color=("#999EA3", "grey30")
            )
        except:
            self.PPStartEntry.configure(
                fg_color=("#e08288", "#6e4441"), border_color="#f51505"
            )
            self.trolleyError = ctk.CTkLabel(
                self.MachineTabs.tab("Trolley Position"),
                text_color="red",
                text="PP Start must be a number",
            )

    def calculateVW(self):
        self.resError.configure(text="")
        self.resError.grid_forget()

        self.density = float(self.densityEntry.get())
        self.elastic_modulus = float(self.elasticEntry.get())
        self.shrinkage = float(self.shrinkageEntry.get())

        self.pipe_length = float(self.pipeLengthEnry.get())
        self.pd = float(self.pipeDiameterEnry.get())
        self.p = float(self.pitchEntry.get()) - float(self.pitchFactorEntry.get())

        self.finalPitchEntry.configure(state="normal")
        self.finalPitchEntry.delete(0, ctk.END)
        self.finalPitchEntry.insert(0, str(self.p))
        self.finalPitchEntry.configure(
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )
        self.cur.execute(
            "SELECT min_wall_thickness FROM diameter WHERE pipe_diameter="
            + str(int(self.pd))
        )
        minWallThickness = self.cur.fetchall()
        self.cur.execute(
            "SELECT mold_diameter FROM diameter WHERE pipe_diameter="
            + str(int(self.pd))
        )
        newPd = self.cur.fetchall()
        try:
            newPd = newPd[0]
        except:
            self.resError.configure(
                text="error: No data for this pipe diameter in diameter"
            )
            self.resError.grid(row=0, column=0, padx=10, pady=5, sticky="w")
            self.pd = [0]

        self.pd = float(newPd[0])

        try:
            minWallThickness = minWallThickness[0]
        except:
            self.resError.configure(
                text="error: No data for this pipe diameter in  wall thickness"
            )
            self.resError.grid(row=0, column=0, padx=10, pady=5, sticky="w")
            minWallThickness = [0]
        minWallThickness = float(minWallThickness[0])
        while minWallThickness > float(self.WallThicknessEntry.get()):
            self.WallThicknessEntry.configure(
                fg_color=("#e08288", "#6e4441"), border_color="#f51505"
            )

            self.WallThicknessEntry.insert(
                0,
                self.Error(
                    self.productionFrame,
                    self.WallThicknessEntry,
                    "Wall thickness",
                    "at least",
                    str(minWallThickness),
                ),
            )
        self.WallThicknessEntry.configure(
            fg_color=("#bababa", "#262626"), border_color=("#999EA3", "grey30")
        )
        self.productionError.configure(text="")
        self.wall_thickness = float(self.WallThicknessEntry.get())
        # pd = 1218.49
        # pipe_length = 6000
        # elastic_modulus = 900
        # density = 0.96
        # shrinkage = 3

        H = (self.wall_thickness * ((100 - self.shrinkage) / 100)) / 2.0
        Sv = 2.0 * H
        J = math.pow(Sv, 3.0) / 12.0
        self.Sn = 1000.0 * J * self.elastic_modulus / math.pow(self.pd, 3.0)
        V1 = self.pd * math.pi * 20.0 * 150.0
        V2 = self.pd * math.pi * 20.0 * 150.0
        V3 = (
            (
                math.pow(
                    self.pd / 2.0
                    + (self.wall_thickness * ((100 - self.shrinkage) / 100)),
                    2.0,
                )
                - math.pow(self.pd / 2.0, 2.0)
            )
            * math.pi
            * self.pipe_length
        )
        W1 = V1 * self.density / 1000000.0
        W2 = V2 * self.density / 1000000.0
        self.W3 = V3 * self.density / 1000000.0
        W = W1 + W2 + self.W3
        self.W0 = W - W2 / 2.0

        self.SnLabel.configure(text="Stiffness: " + str(round(self.Sn, 2)) + " kN/m2")
        self.W1Label.configure(
            text="Socket pipe body weight: " + str(round(W1, 2)) + " kg"
        )
        self.W2Label.configure(
            text="Socket pipe body weight: "
            + str(round(W2, 2))
            + " kg/"
            + str(self.pipe_length / 1000.0)
            + "m"
        )
        self.W3Label.configure(
            text="Flat film tube weight: "
            + str(round(self.W3, 2))
            + " kg/"
            + str(self.pipe_length / 1000.0)
            + "m"
        )

        self.WLabel.configure(
            text="Pipe weight: "
            + str(round(W, 2))
            + " kg/"
            + str(self.pipe_length / 1000.0)
            + "m"
        )
        self.W0Label.configure(
            text="Pipe body weight: "
            + str(round(self.W0, 2))
            + " kg/"
            + str(self.pipe_length / 1000.0)
            + "m"
        )

        self.SnL.configure(text=self.SnLabel.cget("text"))
        self.W0L.configure(text=self.W0Label.cget("text"))

        self.SnL.grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.W0L.grid(row=0, column=1, padx=10, pady=5, sticky="w")
        self.WC.grid(row=1, column=0, padx=10, pady=5, sticky="w")
        self.WP.grid(row=1, column=1, padx=10, pady=5, sticky="w")
        self.moreB.grid(row=2, column=0, padx=10, pady=5, sticky="w")

        self.SnLabel.grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.W1Label.grid(row=1, column=0, padx=10, pady=5, sticky="w")
        self.W2Label.grid(row=2, column=0, padx=10, pady=5, sticky="w")
        self.W3Label.grid(row=0, column=2, padx=10, pady=5, sticky="w")
        self.WLabel.grid(row=1, column=2, padx=10, pady=5, sticky="w")
        self.W0Label.grid(row=2, column=2, padx=10, pady=5, sticky="w")

    def opWLs(self):
        # t1 = threading.Thread(target=self.startls, args=(10,))
        # t2 = threading.Thread(target=self.optimizedPR, args=(10,))

        # t1.start()
        # t2.start()

        # t1.join()
        # t2.join()
        self.resError.grid_forget()
        self.SnL.grid_forget()
        self.W0L.grid_forget()
        self.WC.grid_forget()
        self.WP.grid_forget()
        self.moreB.grid_forget()
        self.moreB.configure(text="more details")
        self.moreFrame.grid_forget()
        self.startls()

    def startls(self, event=None):
        # try:
        #     if self.lp.root.state() == "normal":
        #         try:
        #             self.lp.root.destroy()

        #             # self.lp = ls()
        #             # self.lp.root.mainloop()
        #             # self.lp.root.focus_force()
        #         except:
        #             pass
        #         return

        # except:
        #     pass

        self.ls.grid(
            row=0,
            column=1,
            columnspan=5,
            rowspan=5,
        )
        self.after(200, self.optimizedPR)

        # self.lp = ls()
        # self.lp.root.mainloop()
        # self.lp.root.focus_force()

    def optimizedPR(self, event=None):
        self.SnLabel.configure(text="")
        self.W1Label.configure(text="")
        self.W2Label.configure(text="")
        self.W3Label.configure(text="")
        self.W4Label.configure(text="")
        self.WLabel.configure(text="")
        self.WkLabel.configure(text="")
        self.WpLabel.configure(text="")
        self.WzLabel.configure(text="")
        self.W0Label.configure(text="")
        self.resError.configure(text="")

        while self.pipeLengthEnry.get() == "":
            self.pipeLengthEnry.configure(
                fg_color=("#e08288", "#6e4441"), border_color="#f51505"
            )
            self.productionError.configure(
                text="please enter pipe length",
                fg="#f51505",
            )

        self.pipeLengthEnry.configure(
            fg_color=("white", "grey20"), border_color=("#999EA3", "grey30")
        )

        while self.pipeDiameterEnry.get() == "":
            self.pipeDiameterEnry.configure(
                fg_color=("#e08288", "#6e4441"), border_color="#f51505"
            )

            self.productionError.configure(
                text="please enter pipe diameter",
                fg="#f51505",
            )

        self.pipeDiameterEnry.configure(
            fg_color=("white", "grey20"), border_color=("#999EA3", "grey30")
        )

        while self.pitchFactorEntry.get() == "":
            self.pitchFactorEntry.configure(
                fg_color=("#e08288", "#6e4441"), border_color="#f51505"
            )
            self.productionError.configure(
                text="please enter pitch factor",
                fg="#f51505",
            )

        self.pitchFactorEntry.configure(
            fg_color=("white", "grey20"), border_color=("#999EA3", "grey30")
        )

        while self.reqSnEntries.get() == "":
            self.reqSnEntries.configure(
                fg_color=("#e08288", "#6e4441"), border_color="#f51505"
            )
            self.productionError.configure(
                text="please enter required stiffness",
                fg="#f51505",
            )
        self.productionError.configure(text="")

        self.reqSnEntries.configure(
            fg_color=("white", "grey20"), border_color=("#999EA3", "grey30")
        )

        try:
            if self.isreal(self.densityEntry.get()) == False:
                self.densityEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Density must be a number")
            if float(self.densityEntry.get()) < 0:
                self.densityEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Density must be positive")
            self.densityEntry.configure(
                fg_color=("#bababa", "#262626"), border_color=("#999EA3", "grey30")
            )
            if self.isreal(self.elasticEntry.get()) == False:
                self.elasticEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Elastic Modulus must be a number")
            if float(self.elasticEntry.get()) < 0:
                self.elasticEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Elastic Modulus must be positive")
            self.elasticEntry.configure(
                fg_color=("#bababa", "#262626"), border_color=("#999EA3", "grey30")
            )
            if self.isreal(self.shrinkageEntry.get()) == False:
                self.shrinkageEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Shrinkage must be a number")
            if float(self.shrinkageEntry.get()) < 0:
                self.shrinkageEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Shrinkage must be positive")
            self.shrinkageEntry.configure(
                fg_color=("#bababa", "#262626"), border_color=("#999EA3", "grey30")
            )
            if self.isreal(self.pipeLengthEnry.get()) == False:
                self.pipeLengthEnry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Pipe Lenght must be a number")
            if float(self.pipeLengthEnry.get()) < 0:
                self.pipeLengthEnry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Pipe Lenght must be positive")
            self.pipeLengthEnry.configure(
                fg_color=("white", "grey20"), border_color=("#999EA3", "grey30")
            )
            if self.isreal(self.pipeDiameterEnry.get()) == False:
                self.pipeDiameterEnry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Pipe Diameter must be a number")
            if float(self.pipeDiameterEnry.get()) < 0:
                self.pipeDiameterEnry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Pipe Diameter must be positive")
            self.pipeDiameterEnry.configure(
                fg_color=("white", "grey20"), border_color=("#999EA3", "grey30")
            )
            if self.isreal(self.pitchFactorEntry.get()) == False:
                self.pitchFactorEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Pitch Factor must be a number")
            if float(self.pitchFactorEntry.get()) < 0:
                self.pitchFactorEntry.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Pitch Factor must be positive")
            self.pitchFactorEntry.configure(
                fg_color=("white", "grey20"), border_color=("#999EA3", "grey30")
            )
            if self.isreal(self.reqSnEntries.get()) == False:
                self.reqSnEntries.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Required stiffness must be a number")
            if float(self.reqSnEntries.get()) < 0:
                self.reqSnEntries.configure(
                    fg_color=("#e08288", "#6e4441"), border_color="#f51505"
                )
                raise Exception("Required stiffness must be positive")
            self.reqSnEntries.configure(
                fg_color=("white", "grey20"), border_color=("#999EA3", "grey30")
            )

        except Exception as e:
            self.productionError.configure(text=str(e))
            return

        self.pd = float(self.pipeDiameterEnry.get())
        self.cur.execute(
            "SELECT min_wall_thickness FROM diameter WHERE pipe_diameter="
            + str(int(self.pd))
        )
        minWallThickness = self.cur.fetchall()
        try:
            minWallThickness = minWallThickness[0]
        except:
            self.resError.configure(
                text="error: No data for this pipe diameter in  wall thickness"
            )
            minWallThickness = [0]

        minWallThickness = float(minWallThickness[0])
        flatDies = self.cur.execute(
            "SELECT profile FROM flatDie WHERE thickness >=" + str(minWallThickness)
        ).fetchall()
        claddingDies = self.cur.execute("SELECT profile FROM claddingDie").fetchall()
        for i in range(len(flatDies)):
            flatDies[i] = flatDies[i][0]
        for i in range(len(claddingDies)):
            claddingDies[i] = claddingDies[i][0]
        results = []

        for flatDie in flatDies:
            for claddingDie in claddingDies:
                self.flatDieProfile.set(flatDie)
                self.claddingDieProfile.set(claddingDie)
                self.dieCommand(flatDie)
                self.claddingCommand(claddingDie)
                self.calcPRSn()
                if self.Sn >= float(self.reqSnEntries.get()):
                    self.calcPRW()
                    if self.pp_dist >= 20:
                        self.resError.configure(text="")
                        results.append(
                            [
                                self.flatDieProfile.get(),
                                self.claddingDieProfile.get(),
                                self.W0,
                            ]
                        )
                    else:
                        self.resError.configure(text="Can't be optimized")
                else:
                    self.resError.configure(text="Can't be optimized")
                    # return

        try:
            results.sort(key=lambda x: x[2])

            best = results[0]

        except Exception as e:
            self.SnLabel.configure(text="")
            self.W1Label.configure(text="")
            self.W2Label.configure(text="")
            self.W3Label.configure(text="")
            self.W4Label.configure(text="")
            self.WLabel.configure(text="")
            self.WkLabel.configure(text="")
            self.WpLabel.configure(text="")
            self.WzLabel.configure(text="")
            self.W0Label.configure(text="")
            self.SnL.configure(text="")
            self.W0L.configure(text="")
            self.WC.configure(text="")
            self.WP.configure(text="")

            self.ls.grid_forget()

            self.resError.configure(text="Can't be optimized")
            self.resError.grid(row=0, column=0, padx=10, pady=5, sticky="w")

            return
        self.resError.configure(text="")
        self.resError.grid_forget()
        self.flatDieProfile.set(best[0])
        self.claddingDieProfile.set(best[1])
        self.dieCommand(best[0])
        self.claddingCommand(best[1])
        self.calculatePR()

        self.maxFactorLabel.grid(row=0, column=1, padx=10, pady=5, sticky="w")
        self.maxFactorUnitLabel.grid(row=0, column=3, padx=10, pady=5)
        self.maxFactorEntry.grid(row=0, column=2, padx=10, pady=5, sticky="w")
        self.flat75MaxLabel.grid(row=1, column=1, padx=10, pady=5, sticky="w")
        self.flat75MaxUnitLabel.grid(row=1, column=3, padx=10, pady=5)
        self.flat75MaxEntry.grid(row=1, column=2, padx=10, pady=5, sticky="w")
        self.flat45MaxLabel.grid(row=2, column=1, padx=10, pady=5, sticky="w")
        self.flat45MaxUnitLabel.grid(row=2, column=3, padx=10, pady=5)
        self.flat45MaxEntry.grid(row=2, column=2, padx=10, pady=5, sticky="w")
        self.cladding75MaxLabel.grid(row=3, column=1, padx=10, pady=5, sticky="w")
        self.cladding75MaxUnitLabel.grid(row=3, column=3, padx=10, pady=5)
        self.cladding75MaxEntry.grid(row=3, column=2, padx=10, pady=5, sticky="w")
        self.cladding45MaxLabel.grid(row=4, column=1, padx=10, pady=5, sticky="w")
        self.cladding45MaxUnitLabel.grid(row=4, column=3, padx=10, pady=5)
        self.cladding45MaxEntry.grid(row=4, column=2, padx=10, pady=5, sticky="w")
        self.FlatExtruder75Label.grid(row=1, column=4, padx=10, pady=5, sticky="w")
        self.FlatExtruder75UnitLabel.grid(row=1, column=6, padx=10, pady=5)
        self.FlatExtruder75Entry.grid(row=1, column=5, padx=10, pady=5, sticky="w")
        self.FlatExtruder45Label.grid(row=2, column=4, padx=10, pady=5, sticky="w")
        self.FlatExtruder45UnitLabel.grid(row=2, column=6, padx=10, pady=5)
        self.FlatExtruder45Entry.grid(row=2, column=5, padx=10, pady=5, sticky="w")
        self.CladdingExtruder75Label.grid(row=3, column=4, padx=10, pady=5, sticky="w")
        self.CladdingExtruder75UnitLabel.grid(row=3, column=6, padx=10, pady=5)
        self.CladdingExtruder75Entry.grid(row=3, column=5, padx=10, pady=5, sticky="w")
        self.CladdingExtruder45Label.grid(row=4, column=4, padx=10, pady=5, sticky="w")
        self.CladdingExtruder45UnitLabel.grid(row=4, column=6, padx=10, pady=5)
        self.CladdingExtruder45Entry.grid(row=4, column=5, padx=10, pady=5, sticky="w")

        self.FlatExtruder75LabelF.grid(row=1, column=7, padx=10, pady=5, sticky="w")
        self.FlatExtruder75UnitLabelF.grid(row=1, column=9, padx=10, pady=5)
        self.FlatExtruder75EntryF.grid(row=1, column=8, padx=10, pady=5, sticky="w")
        self.FlatExtruder45LabelF.grid(row=2, column=7, padx=10, pady=5, sticky="w")
        self.FlatExtruder45UnitLabelF.grid(row=2, column=9, padx=10, pady=5)
        self.FlatExtruder45EntryF.grid(row=2, column=8, padx=10, pady=5, sticky="w")
        self.CladdingExtruder75LabelF.grid(row=3, column=7, padx=10, pady=5, sticky="w")
        self.CladdingExtruder75UnitLabelF.grid(row=3, column=9, padx=10, pady=5)
        self.CladdingExtruder75EntryF.grid(row=3, column=8, padx=10, pady=5, sticky="w")
        self.CladdingExtruder45LabelF.grid(row=4, column=7, padx=10, pady=5, sticky="w")
        self.CladdingExtruder45UnitLabelF.grid(row=4, column=9, padx=10, pady=5)
        self.CladdingExtruder45EntryF.grid(row=4, column=8, padx=10, pady=5, sticky="w")

        self.MoldSpeedLabel.grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.MoldSpeedUnitLabel.grid(row=0, column=2, padx=10, pady=5)
        self.MoldSpeedEntry.grid(row=0, column=1, padx=10, pady=5, sticky="w")
        self.ppSpeedLabel.grid(row=1, column=3, padx=10, pady=5, sticky="w")
        self.ppSpeedUnitLabel.grid(row=1, column=5, padx=10, pady=5)
        self.ppSpeedEntry.grid(row=1, column=4, padx=10, pady=5, sticky="w")
        self.trolleySpeedLabel.grid(row=0, column=3, padx=10, pady=5, sticky="w")
        self.trolleySpeedUnitLabel.grid(row=0, column=5, padx=10, pady=5)
        self.trolleySpeedEntry.grid(row=0, column=4, padx=10, pady=5, sticky="w")
        self.ppSpeedFractionLabel.grid(row=1, column=0, padx=10, pady=5, sticky="w")
        self.ppSpeedFractionUnitLabel.grid(row=1, column=2, padx=10, pady=5)
        self.ppSpeedFractionEntry.grid(row=1, column=1, padx=10, pady=5, sticky="w")

        self.MouthStartLabel.grid(
            row=0,
            column=0,
            padx=10,
            pady=5,
        )
        self.MouthEndLabel.grid(
            row=0,
            column=2,
            padx=10,
            pady=5,
        )
        self.CarriageReturnPositionLabel.grid(
            row=0,
            column=4,
            padx=10,
            pady=5,
        )
        self.CarriageRotationsLabel.grid(
            row=0,
            column=6,
            padx=10,
            pady=5,
        )
        self.CarriageReturnDelayLabel.grid(
            row=0,
            column=8,
            padx=10,
            pady=5,
        )
        self.PPStartLabel.grid(
            row=1,
            column=0,
            padx=10,
            pady=5,
        )
        self.PPEndLabel.grid(
            row=1,
            column=2,
            padx=10,
            pady=5,
        )
        self.SocketStartLabel.grid(
            row=1,
            column=4,
            padx=10,
            pady=5,
        )
        self.SocketEndLabel.grid(
            row=1,
            column=6,
            padx=10,
            pady=5,
        )
        self.pitchAngleLabel.grid(
            row=2,
            column=0,
            padx=10,
            pady=5,
        )
        self.trolleyRotataionsLabel.grid(
            row=2,
            column=0,
            padx=10,
            pady=5,
        )
        self.trolleyIncreaseTimeLabel.grid(
            row=2,
            column=2,
            padx=10,
            pady=5,
        )
        self.ppExtruderStartLabel.grid(
            row=2,
            column=4,
            padx=10,
            pady=5,
        )
        self.ppEndDelayLabel.grid(
            row=2,
            column=6,
            padx=10,
            pady=5,
        )

        self.MouthStartEntry.grid(
            row=0,
            column=1,
            padx=10,
            pady=5,
        )
        self.MouthStartEntry.delete(0, ctk.END)
        self.MouthStartEntry.insert(0, "115")
        self.MouthEndEntry.grid(row=0, column=3, padx=10, pady=5)
        self.MouthEndEntry.delete(0, ctk.END)
        self.MouthEndEntry.insert(0, "680")
        self.CarriageReturnPositionEntry.delete(0, ctk.END)
        self.CarriageReturnPositionEntry.insert(0, "350")
        self.CarriageReturnPositionEntry.grid(
            row=0,
            column=5,
            padx=10,
            pady=5,
        )
        self.CarriageRotationsEntry.delete(0, ctk.END)
        self.CarriageRotationsEntry.insert(0, "1")
        self.CarriageRotationsEntry.grid(
            row=0,
            column=7,
            padx=10,
            pady=5,
        )

        self.CarriageRotationsEntry.bind("<Return>", self.calcSpeeds)
        self.CarriageReturnDelayEntry.grid(
            row=0,
            column=9,
            padx=10,
            pady=5,
        )
        self.PPStartEntry.grid(
            row=1,
            column=1,
            padx=10,
            pady=5,
        )

        self.PPStartEntry.bind("<Return>", self.calcPPExSt)
        self.PPStartEntry.delete(0, ctk.END)
        self.PPStartEntry.insert(0, "680")
        self.PPEndEntry.grid(
            row=1,
            column=3,
            padx=10,
            pady=5,
        )
        self.PPEndEntry.delete(0, ctk.END)
        self.PPEndEntry.insert(0, "6114")
        self.SocketStartEntry.grid(
            row=1,
            column=5,
            padx=10,
            pady=5,
        )
        self.SocketStartEntry.delete(0, ctk.END)
        self.SocketStartEntry.insert(0, "6000")
        self.SocketEndEntry.grid(row=1, column=7, padx=10, pady=5)
        self.SocketEndEntry.delete(0, ctk.END)
        self.SocketEndEntry.insert(0, "6115")
        temp = math.degrees(
            math.atan((self.p / 4) / (self.pd / 2 + self.s1 + self.ppd + self.s4 * 2))
        )
        self.pitchAngleEntry.configure(state="normal")
        self.pitchAngleEntry.delete(0, ctk.END)
        self.pitchAngleEntry.insert(0, str(round(temp, 2)))
        self.pitchAngleEntry.grid(row=2, column=1, padx=10, pady=5)
        self.pitchAngleEntry.configure(state="disabled")
        self.trolleyRotataionsEntry.grid(
            row=2,
            column=1,
            padx=10,
            pady=5,
        )
        self.trolleyRotataionsEntry.delete(0, ctk.END)
        self.trolleyRotataionsEntry.insert(0, "1")

        self.trolleyRotataionsEntry.bind("<Return>", self.calcSpeeds)
        self.trolleyIncreaseTimeEntry.grid(
            row=2,
            column=3,
            padx=10,
            pady=5,
        )

        self.trolleyIncreaseTimeEntry.bind("<Return>", self.calcTrolley)
        self.ppExtruderStartEntry.grid(
            row=2,
            column=5,
            padx=10,
            pady=5,
        )
        self.ppEndDelayEntry.grid(
            row=2,
            column=7,
            padx=10,
            pady=5,
        )
        self.f45(None)
        self.c45(None)
        self.calcSpeeds(None)
        self.calcPPExSt(None)
        try:
            self.after(300, self.ls.grid_forget())
            pass
        except:
            pass

    def calcTrolley(self, event):
        temp = (self.p * float(self.trolleyIncreaseTimeEntry.get())) / 250
        self.ppEndDelayEntry.configure(state="normal")
        self.ppEndDelayEntry.delete(0, ctk.END)
        self.ppEndDelayEntry.insert(0, str(round(temp, 2)))
        self.ppEndDelayEntry.configure(
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )
        self.trolleyIncreaseTimeEntry.configure(
            fg_color=("#bababa", "#262626"), border_color=("#999EA3", "grey30")
        )

    def calcPRSn(self):
        self.resError.configure(text="")
        self.resError.grid_forget()
        self.density = float(self.densityEntry.get())
        self.elastic_modulus = float(self.elasticEntry.get())
        self.shrinkage = float(self.shrinkageEntry.get())

        self.pipe_length = float(self.pipeLengthEnry.get())
        self.pd = float(self.pipeDiameterEnry.get())
        self.p = float(self.pitchEntry.get()) - float(self.pitchFactorEntry.get())
        self.finalPitchEntry.configure(state="normal")
        self.finalPitchEntry.delete(0, ctk.END)
        self.finalPitchEntry.insert(0, str(self.p))
        self.finalPitchEntry.configure(
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )
        self.cur.execute(
            "SELECT min_wall_thickness FROM diameter WHERE pipe_diameter="
            + str(int(self.pd))
        )
        minWallThickness = self.cur.fetchall()
        self.cur.execute(
            "SELECT mold_diameter FROM diameter WHERE pipe_diameter="
            + str(int(self.pd))
        )
        newPd = self.cur.fetchall()
        try:
            newPd = newPd[0]
        except:
            self.resError.configure(
                text="error: No data for this pipe diameter in diameter"
            )
            self.resError.grid(row=0, column=0, padx=10, pady=5, sticky="w")
            return

        self.pd = float(newPd[0])

        try:
            minWallThickness = minWallThickness[0]
        except:
            self.resError.configure(
                text="error: No data for this pipe diameter in  wall thickness"
            )
            self.resError.grid(row=0, column=0, padx=10, pady=5, sticky="w")
            return

        minWallThickness = float(minWallThickness[0])

        while minWallThickness > float(self.WallThicknessEntry.get()):
            self.WallThicknessEntry.configure(
                fg_color=("#e08288", "#6e4441"), border_color="#f51505"
            )

            self.WallThicknessEntry.insert(
                0,
                self.Error(
                    self.productionFrame,
                    self.WallThicknessEntry,
                    "Wall thickness",
                    "at least",
                    str(minWallThickness),
                ),
            )
        self.productionError.configure(text="")

        self.WallThicknessEntry.configure(
            fg_color=("#bababa", "#262626"), border_color=("#999EA3", "grey30")
        )

        self.s1 = (
            float(self.WallThicknessEntry.get()) * float(100 - self.shrinkage) / 100.0
        )
        self.ppd = float(self.PPDiameterEntry.get())
        self.s4 = (
            float(self.PPFilmThicknessEntry.get()) * float(100 - self.shrinkage) / 100.0
        )

        Y1 = self.s1 / 2.0
        Y2 = self.s1 + self.s4 + 0.9 * self.ppd / 2.0
        A1 = self.s1 * self.p
        A2 = (0.9 * self.ppd / 2.0 + self.s4) ** 2 * math.pi - (
            0.9 * self.ppd / 2.0
        ) ** 2 * math.pi
        H = (Y1 * A1 + Y2 * A2) / (A1 + A2)
        Y11 = abs(H - Y1)
        Y21 = abs(Y2 - H)
        J1 = self.p * (self.s1) ** 3 / 12.0
        J2 = (
            math.pi
            / 64.0
            * ((0.9 * self.ppd + 2.0 * self.s4) ** 4 - (0.9 * self.ppd) ** 4)
        )
        J = (J1 + A1 * Y11**2 + J2 + A2 * Y21**2) / self.p
        self.Sn = 1000.0 * J * self.elastic_modulus / self.pd**3

        self.SnLabel.configure(text="Stiffness: " + str(round(self.Sn, 2)) + " kN/m2")
        self.SnL.configure(text=self.SnLabel.cget("text"))

        self.SnL.grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.SnLabel.grid(row=0, column=0, padx=10, pady=5, sticky="w")

    def calcPRW(self):
        self.resError.configure(text="")
        self.resError.grid_forget()
        self.density = float(self.densityEntry.get())
        self.elastic_modulus = float(self.elasticEntry.get())
        self.shrinkage = float(self.shrinkageEntry.get())

        self.pipe_length = float(self.pipeLengthEnry.get())
        self.pd = float(self.pipeDiameterEnry.get())
        self.p = float(self.pitchEntry.get()) - float(self.pitchFactorEntry.get())
        self.finalPitchEntry.configure(state="normal")
        self.finalPitchEntry.delete(0, ctk.END)
        self.finalPitchEntry.insert(0, str(self.p))
        self.finalPitchEntry.configure(
            state="disabled",
            fg_color=("#bababa", "#262626"),
        )
        self.cur.execute(
            "SELECT min_wall_thickness FROM diameter WHERE pipe_diameter="
            + str(int(self.pd))
        )
        minWallThickness = self.cur.fetchall()
        self.cur.execute(
            "SELECT mold_diameter FROM diameter WHERE pipe_diameter="
            + str(int(self.pd))
        )
        newPd = self.cur.fetchall()
        try:
            newPd = newPd[0]
        except:
            self.resError.configure(
                text="error: No data for this pipe diameter in diameter"
            )
            self.resError.grid(row=0, column=0, padx=10, pady=5, sticky="w")
            return

        self.pd = float(newPd[0])

        try:
            minWallThickness = minWallThickness[0]
        except:
            self.resError.configure(
                text="error: No data for this pipe diameter in  wall thickness"
            )
            self.resError.grid(row=0, column=0, padx=10, pady=5, sticky="w")
            return

        minWallThickness = float(minWallThickness[0])

        while minWallThickness > float(self.WallThicknessEntry.get()):
            self.WallThicknessEntry.configure(
                fg_color=("#e08288", "#6e4441"), border_color="#f51505"
            )

            self.WallThicknessEntry.insert(
                0,
                self.Error(
                    self.productionFrame,
                    self.WallThicknessEntry,
                    "Wall thickness",
                    "at least",
                    str(minWallThickness),
                ),
            )
        self.productionError.configure(text="")

        self.WallThicknessEntry.configure(
            fg_color=("#bababa", "#262626"), border_color=("#999EA3", "grey30")
        )

        self.s1 = (
            float(self.WallThicknessEntry.get()) * float(100 - self.shrinkage) / 100.0
        )
        self.ppd = float(self.PPDiameterEntry.get())
        self.s4 = (
            float(self.PPFilmThicknessEntry.get()) * float(100 - self.shrinkage) / 100.0
        )

        # Equations for Pipe Weight
        V1 = self.pd * math.pi * 20.0 * 150.0
        V2 = self.pd * math.pi * 20.0 * 150.0
        V3 = (
            ((self.pd / 2.0 + self.s1) ** 2 - (self.pd / 2.0) ** 2)
            * math.pi
            * self.pipe_length
        )
        Sbf = ((self.ppd / 2.0 + self.s4) ** 2 - (self.ppd / 2.0) ** 2) * math.pi
        Lpp = self.pipe_length / self.p * (self.pd + self.s1 * 2) * math.pi
        V4 = Sbf * Lpp
        W1 = V1 * self.density / 1000000.0
        W2 = V2 * self.density / 1000000.0
        self.W3 = V3 * self.density / 1000000.0
        self.W4 = V4 * self.density / 1000000.0
        W = W1 + W2 + self.W3 + self.W4
        self.cur.execute(
            "SELECT weight FROM ppwt WHERE pp_diameter=" + str(int(self.ppd))
        )
        tempWk = self.cur.fetchall()
        try:
            tempWk = tempWk[0]
        except:
            tempWk = [0]
        tempWk = float(tempWk[0])
        Wk = tempWk / 1000.0
        Wp = Lpp * Wk / 1000.0
        Wz = W1 + W2 + self.W3 + self.W4 + Wp
        self.W0 = Wz - W2 / 2.0
        self.pp_dist = self.p - self.ppd - self.s4 * 2.0

        if self.pp_dist < 20:
            self.resError.configure(
                text="pp_dist can't be less than 20",
            )
            self.resError.grid(row=0, column=0, padx=10, pady=5, sticky="w")
            return
        self.resError.configure(text="")
        self.resError.grid_forget()

        self.W1Label.configure(
            text="Socket pipe body weight: " + str(round(W1, 2)) + " kg"
        )
        self.W2Label.configure(
            text="Socket pipe body weight: "
            + str(round(W2, 2))
            + " kg/"
            + str(self.pipe_length / 1000.0)
            + "m"
        )
        self.W3Label.configure(
            text="Flat film tube weight: "
            + str(round(self.W3, 2))
            + " kg/"
            + str(self.pipe_length / 1000.0)
            + "m"
        )
        self.W4Label.configure(
            text="Coated film body weight: "
            + str(round(self.W4, 2))
            + " kg/"
            + str(self.pipe_length / 1000.0)
            + "m"
        )
        self.WLabel.configure(
            text="Pipe weight excluding pp pipe: "
            + str(round(W, 2))
            + " kg/"
            + str(self.pipe_length / 1000.0)
            + "m"
        )
        self.WkLabel.configure(text="PP pipe weight: " + str(round(Wk, 2)) + " kg/m")
        self.WpLabel.configure(
            text="PP pipe weight: "
            + str(round(Wp, 2))
            + " kg/"
            + str(self.pipe_length / 1000.0)
            + "m"
        )
        self.WzLabel.configure(
            text="Pipe weight: "
            + str(round(Wz, 2))
            + " kg/"
            + str(self.pipe_length / 1000.0)
            + "m"
        )
        self.W0Label.configure(
            text="Pipe body weight: "
            + str(round(self.W0, 2))
            + " kg/"
            + str(self.pipe_length / 1000.0)
            + "m"
        )

        self.W0L.configure(text=self.W0Label.cget("text"))

        wp = str(self.WpLabel.cget("text")).split(": ")[1]
        wpt = ""
        for i in range(len(wp)):
            if wp[i].isalpha() is False:
                wpt = wpt + wp[i]
            else:
                break
        w0 = str(self.W0Label.cget("text")).split(": ")[1]
        w0t = ""
        for i in range(len(w0)):
            if w0[i].isalpha() is False:
                w0t = w0t + w0[i]
            else:
                break
        temp = (float(wpt) / float(w0t)) * 100
        self.WC.configure(text="pp weight percentage: " + str(round(temp, 2)) + "%")
        self.WP.configure(
            text="pe weight percentage: " + str(round(100 - temp, 2)) + "%"
        )

        self.SnL.grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.W0L.grid(row=0, column=1, padx=10, pady=5, sticky="w")
        self.WC.grid(row=1, column=0, padx=10, pady=5, sticky="w")
        self.WP.grid(row=1, column=1, padx=10, pady=5, sticky="w")
        self.moreB.grid(row=2, column=0, padx=10, pady=5, sticky="w")

        self.W1Label.grid(row=1, column=0, padx=10, pady=5, sticky="w")
        self.W2Label.grid(row=2, column=0, padx=10, pady=5, sticky="w")
        self.W3Label.grid(row=3, column=0, padx=10, pady=5, sticky="w")
        self.W4Label.grid(row=4, column=0, padx=10, pady=5, sticky="w")
        self.WLabel.grid(row=0, column=2, padx=10, pady=5, sticky="w")
        self.WkLabel.grid(row=1, column=2, padx=10, pady=5, sticky="w")
        self.WpLabel.configure(
            text=self.WpLabel.cget("text") + "(" + str(round(temp, 2)) + "%)"
        )
        self.WpLabel.grid(row=2, column=2, padx=10, pady=5, sticky="w")
        self.WzLabel.grid(row=3, column=2, padx=10, pady=5, sticky="w")
        self.W0Label.grid(row=4, column=2, padx=10, pady=5, sticky="w")

    def calculatePR(self):
        self.calcPRSn()
        self.calcPRW()

    def setResize(self, flag):
        self.resizable(flag, flag)

    def edit(self, table=str, create=False, row=-1, id=-1):
        try:
            if self.editW.state() == "normal":
                try:
                    self.editW.destroy()
                    self.editW = Edit(table, create, row, id, self)
                except:
                    pass
                return
            else:
                self.refresh(table, None)
        except:
            pass

        self.editW = Edit(table, create, row, id, self)

        self.refresh(table, None)

    def delete(self, table, row):
        dialog = ctk.CTkInputDialog(
            text='Type "DELETE" to delete this row:', title="Delete"
        )
        dialog.geometry(
            "300x200+"
            + str(int((self.winfo_screenwidth() - 300) / 2))
            + "+"
            + str(int((self.winfo_screenheight() - 200) / 2))
        )
        if dialog.get_input() == "DELETE":
            self.cur.execute("DELETE FROM " + table + " WHERE id=" + str(row))
            self.conn.commit()
            self.select_frame_by_name(table)

    def refresh(self, table, event):
        self.Frame.grid_forget()
        self.Frame.destroy()

        self.cur.execute("SELECT * FROM " + table)
        # self.Frame.destroy()
        # self.Frame.grid
        self.Frame = ctk.CTkScrollableFrame(
            self,
            orientation="both",
            scrollbar_button_color="grey30",
            bg_color="transparent",
            fg_color="transparent",
        )
        self.Frame.grid_columnconfigure(0, weight=1)
        self.Frame.grid_columnconfigure(1, weight=1)

        self.Headers = [description[0] for description in self.cur.description[1::]]
        self.headers = self.Headers.copy()
        for i in range(len(self.Headers)):
            if self.headers[i] == "profile":
                pass
            else:
                self.headers[i] = self.headers[i].replace("_", " ")
        self.RowsData = self.cur.fetchall()
        self.NumRows = len(self.RowsData)
        self.Entries = []

        if len(self.Headers) > 0:
            create = ctk.CTkButton(
                self.Frame,
                text="Add",
                command=lambda: self.edit(table, True, -1, -1),
                width=15,
                fg_color="#5696b0",
                # fg_color="transparent",
            )
            create.grid(row=0, column=0, padx=10, pady=5)
            for col, header in enumerate(self.headers):
                self.label = ctk.CTkLabel(self.Frame, text=header)
                self.label.grid(row=0, column=col + 1, padx=10, pady=5)

        if len(self.RowsData) > 0:
            for row, row_data in enumerate(self.RowsData, start=1):
                edit = ctk.CTkButton(
                    self.Frame,
                    text="",
                    command=lambda row=row, temp=row_data[0]: self.edit(
                        table, False, row, int(temp)
                    ),
                    width=15,
                    height=15,
                    image=ctk.CTkImage(
                        light_image=Image.open("icons/edit.png"),
                        size=(15, 15),
                        dark_image=Image.open("icons/editD.png"),
                    ),
                    bg_color="transparent",
                    fg_color="transparent",
                    hover_color="#c4c4c4",
                )
                edit.grid(row=row, column=0, padx=10, pady=5)
                delete = ctk.CTkButton(
                    self.Frame,
                    text="",
                    command=lambda temp=row_data[0]: self.delete(table, int(temp)),
                    width=15,
                    height=15,
                    image=ctk.CTkImage(
                        light_image=Image.open("icons/delete.png"),
                        size=(15, 15),
                        dark_image=Image.open("icons/deleteD.png"),
                    ),
                    bg_color="transparent",
                    fg_color="transparent",
                    hover_color="#c4c4c4",
                )
                delete.grid(
                    row=row, column=len(self.Headers) + 1, padx=10, pady=5, sticky="w"
                )
                for col, value in enumerate(row_data[1::]):
                    entry = ctk.CTkEntry(self.Frame, width=100, state="normal")
                    if value is not None:
                        entry.insert(ctk.END, value)
                    entry.configure(
                        state="disabled",
                        fg_color=("#bababa", "#262626"),
                    )
                    entry.grid(row=row, column=col + 1, padx=10, pady=5)
                    self.Entries.append(entry)


main = main()
main.mainloop()
